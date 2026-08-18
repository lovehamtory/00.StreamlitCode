from __future__ import annotations

import re
from concurrent.futures import FIRST_COMPLETED, Future, ThreadPoolExecutor, wait
from contextlib import ExitStack
from dataclasses import dataclass
from datetime import datetime
from html import escape
from io import BytesIO
from typing import Any

import oracledb
import pandas as pd
import streamlit as st

try:
    import pyodbc
except ImportError:
    pyodbc = None


MAPPING_TABLE = "TB_MIG_TABLE_INFO"
TARGET_OWNER = "PCERP_RENTALAPP_MIG"
SOURCE_DATABASES = ("CERDB", "INSIDEBANK", "JBNDB", "MEMDB", "PREEDDB", "SALDB")
RESULT_COLUMNS = [
    "소스DB",
    "소스TABLE",
    "소스ENTITY",
    "타겟TABLE",
    "검증결과",
    "오류여부",
    "소스CNT",
    "타겟CNT",
    "차이",
    "소스DATA MB",
    "소스DATA GB",
    "소스INDEX MB",
    "소스합계 MB",
    "비고",
]


@dataclass(frozen=True)
class MappingRow:
    source_db: str
    source_table: str
    source_entity: str
    target_table: str


def init_state() -> None:
    defaults: dict[str, Any] = {
        "countcheck_results": pd.DataFrame(columns=RESULT_COLUMNS),
        "countcheck_summary": None,
        "countcheck_mappings": pd.DataFrame(),
        "countcheck_progress": None,
    }
    for key, value in defaults.items():
        st.session_state.setdefault(key, value)


def clean_text(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def upper_text(value: object) -> str:
    return re.sub(r"\s+", " ", clean_text(value)).upper()


def oracle_identifier(value: str) -> str:
    raw = clean_text(value)
    if raw.startswith('"') and raw.endswith('"'):
        raw = raw[1:-1].replace('""', '"')
    name = upper_text(raw)
    if not name or "." in name or any(ord(character) < 32 for character in name) or len(name.encode("utf-8")) > 128:
        raise ValueError(f"Oracle 식별자 형식이 올바르지 않습니다: {value}")
    return f'"{name.replace(chr(34), chr(34) * 2)}"'


def mssql_identifier(value: str) -> str:
    parts = [part.strip() for part in clean_text(value).split(".")]
    normalized_parts: list[str] = []
    for part in parts:
        if part.startswith("[") and part.endswith("]"):
            part = part[1:-1].replace("]]", "]")
        if not part or len(part) > 128 or any(ord(character) < 32 for character in part):
            raise ValueError(f"MSSQL 식별자 형식이 올바르지 않습니다: {value}")
        normalized_parts.append(part)
    if not normalized_parts:
        raise ValueError(f"MSSQL 식별자 형식이 올바르지 않습니다: {value}")
    return ".".join(f"[{part.replace(']', ']]')}]" for part in normalized_parts)


def target_table_name(value: str) -> str:
    parts = [part.strip().replace('"', "") for part in clean_text(value).split(".") if part.strip()]
    if not parts:
        raise ValueError("타겟 테이블명이 없습니다.")
    return parts[-1]


def derived_target_table(source_db: str, source_table: str) -> str:
    raw_name = re.sub(r"[^A-Z0-9_$#]", "_", upper_text(target_table_name(source_table)))
    return f"{upper_text(source_db)}_{raw_name}".strip("_")


def now_label() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def duration_label(seconds: float) -> str:
    total_seconds = max(int(seconds), 0)
    hours, remainder = divmod(total_seconds, 3_600)
    minutes, seconds_part = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds_part:02d}"


def elapsed_label(started: object, finished: object) -> str:
    try:
        started_at = datetime.strptime(clean_text(started), "%Y-%m-%d %H:%M:%S")
        finished_text = clean_text(finished)
        finished_at = datetime.strptime(finished_text, "%Y-%m-%d %H:%M:%S") if finished_text and finished_text != "-" else datetime.now()
        return duration_label((finished_at - started_at).total_seconds())
    except ValueError:
        return "-"


def time_label(value: object) -> str:
    raw_value = clean_text(value)
    matched = re.search(r"(\d{2}:\d{2}:\d{2})$", raw_value)
    return matched.group(1) if matched else (raw_value or "-")


def empty_result_frame() -> pd.DataFrame:
    return pd.DataFrame(columns=RESULT_COLUMNS)


def apply_operations_style() -> None:
    st.markdown(
        """
        <style>
        [data-testid="stAppViewContainer"] { background: radial-gradient(circle at 18% -10%, #17385f 0, #0a1324 34%, #07101e 100%); }
        [data-testid="stHeader"] { background: transparent; }
        [data-testid="stSidebar"] > div:first-child { background: #0b1729; border-right: 1px solid #203a5a; }
        [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p { color: #d9e7f7; }
        .block-container { max-width: 1520px; padding-top: 2.1rem; padding-bottom: 3.5rem; }
        .ops-header { padding: 1.35rem 1.55rem 1.2rem; margin-bottom: 1cm; background: linear-gradient(105deg, rgba(17, 43, 76, .98), rgba(10, 22, 39, .78)); border: 1px solid #274c77; border-left: 5px solid #35a8ff; border-radius: 14px; box-shadow: 0 18px 44px rgba(0, 0, 0, .22); }
        .ops-kicker { color: #75c7ff; font-size: .77rem; font-weight: 700; letter-spacing: .13em; }
        .ops-header h1 { margin: .22rem 0 .38rem; color: #f5f9ff; font-size: 2rem; line-height: 1.1; letter-spacing: -.025em; }
        .ops-header p { margin: 0; color: #9fb5ce; font-size: .9rem; }
        .ops-card { height: 102px; box-sizing: border-box; display: flex; flex-direction: column; justify-content: center; gap: 14px; padding: 16px; background: rgba(16, 33, 56, .9); border: 1px solid #254d73; border-radius: 10px; }
        .ops-card-label { color: #c9dcef; font-size: 16px; font-weight: 720; line-height: 1; letter-spacing: -.01em; }
        .ops-card-value { color: #f4f8fd; font-size: 16px; font-weight: 780; line-height: 1; letter-spacing: -.01em; white-space: nowrap; }
        .ops-section-title { display: flex; align-items: center; gap: 7px; padding-bottom: 0; color: #f1f7ff; font-size: 16px; font-weight: 780; line-height: 1; }
        .ops-section-icon { color: #6fc4ff; font-size: 16px; }
        .ops-system-title { color: #dceeff; font-size: 14px; font-weight: 760; line-height: 1; }
        .ops-progress-summary { color: #d5e3f1; font-size: 16px; font-weight: 660; line-height: 1.45; white-space: nowrap; }
        .ops-progress-detail { color: #d5e3f1; font-size: 14px; font-weight: 640; line-height: 1.45; white-space: nowrap; }
        .ops-progress-list { display: grid; gap: .42rem; }
        .ops-progress-name { color: #eef7ff; font-weight: 760; }
        .ops-progress-count { color: #d5e9ff; font-weight: 720; }
        .ops-progress-time { color: #9db5cc; }
        .ops-state { display: inline-block; border-radius: 4px; padding: 2px 5px; font-size: 12px; font-weight: 760; line-height: 1.1; }
        .ops-state-done { color: #69e393; background: rgba(49, 154, 89, .18); }
        .ops-state-running { color: #ffc466; background: rgba(182, 116, 28, .2); }
        .ops-state-error { color: #ff8d97; background: rgba(185, 51, 64, .2); }
        .ops-error-clear { color: #93aabe; }
        .ops-error-active { color: #ff8d97; font-weight: 720; }
        [data-testid="stMetric"] { background: rgba(16, 33, 56, .9); border: 1px solid #254566; border-radius: 10px; padding: .8rem .95rem; }
        [data-testid="stMetricLabel"] { color: #a9bed5; font-size: .76rem; letter-spacing: .015em; }
        [data-testid="stMetricValue"] { color: #f4f8fd; font-size: 1.12rem; }
        [data-testid="stDataFrame"] { border: 1px solid #284666; border-radius: 10px; overflow: hidden; }
        [data-testid="stDataFrame"] [role="columnheader"] { background: #142b48; color: #dcebfa; }
        [data-testid="stProgressBar"] > div > div { background: linear-gradient(90deg, #118ddd, #41b8ff); }
        [data-testid="stButton"] > button { border: 1px solid #2f78ad; background: #123d66; color: #f4f9ff; border-radius: 8px; font-weight: 650; }
        [data-testid="stButton"] > button[kind="primary"] { background: linear-gradient(100deg, #057bc5, #169fe9); border-color: #42b9ff; }
        [data-testid="stSelectbox"] > div > div, [data-testid="stTextInput"] input, [data-testid="stFileUploader"] section { background: #0e2139; border-color: #315273; color: #eff7ff; }
        [data-testid="stExpander"] { border-color: #284a6c; background: #0d1c30; }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_operations_header() -> None:
    st.markdown(
        """
        <div class="ops-header">
          <div class="ops-kicker">DATA MIGRATION CONTROL ROOM</div>
          <h1>▣ MSSQL ↔ ORACLE 건수 검증</h1>
          <p>⚙️ Created by ♡홍율파파</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def snapshot_card(label: str, value: str) -> str:
    return (
        f'<div class="ops-card"><div class="ops-card-label">{label}</div>'
        f'<div class="ops-card-value">{value}</div></div>'
    )


def render_card_row(cards: tuple[tuple[str, str], ...]) -> None:
    for column, card in zip(st.columns(6, gap="small"), cards):
        with column:
            st.markdown(snapshot_card(card[0], card[1]), unsafe_allow_html=True)


def render_section_title(icon: str, title: str, container: Any | None = None) -> None:
    target = st if container is None else container
    target.markdown(
        f'<div class="ops-section-title"><span class="ops-section-icon">{icon}</span><span>{title}</span></div>',
        unsafe_allow_html=True,
    )


def render_fixed_gap(pixels: int, container: Any | None = None) -> None:
    target = st if container is None else container
    target.space(pixels)


def render_operations_snapshot(summary: dict[str, Any] | None, target_count: int = 0) -> None:
    mssql = (summary or {}).get("mssql", {})
    oracle = (summary or {}).get("oracle", {})
    mismatch_count_value = int(mssql.get("mismatches", 0))
    error_count_value = int(mssql.get("errors", 0)) + int(oracle.get("errors", 0))
    finished_values = [value for value in (mssql.get("finished"), oracle.get("finished")) if value and value != "-"]
    elapsed_values = [value for value in (mssql.get("elapsed"), oracle.get("elapsed")) if value]
    finished = max(finished_values) if finished_values else "-"
    elapsed = max(elapsed_values) if elapsed_values else "-"
    cards = (
        ("▣ 대상 테이블", f"{target_count:,}건"),
        ("◷ 조회 시작", mssql.get("started", "-")),
        ("◷ 조회 종료", finished),
        ("◴ 소요 시간", elapsed),
        ("⇄ 불일치", f"{mismatch_count_value:,}건"),
        ("⚠ 오류", f"{error_count_value:,}건"),
    )
    render_card_row(cards)


@st.cache_resource(show_spinner=False)
def get_oracle_pool(
    host: str,
    port: int,
    service_name: str,
    user: str,
    password: str,
    max_connections: int,
) -> oracledb.ConnectionPool:
    return oracledb.create_pool(
        user=user,
        password=password,
        dsn=f"{host}:{port}/{service_name}",
        min=1,
        max=max_connections,
        increment=1,
    )


def oracle_config() -> dict[str, Any]:
    required = ("user", "password", "host", "port", "service_name", "owner")
    if "oracle" not in st.secrets:
        raise ValueError(".streamlit/secrets.toml에 [oracle] 설정이 없습니다.")
    config = dict(st.secrets["oracle"])
    missing = [key for key in required if not clean_text(config.get(key))]
    if missing:
        raise ValueError(f"[oracle] 필수 항목이 없습니다: {', '.join(missing)}")
    return config


def mssql_config() -> dict[str, Any]:
    required = ("host", "port", "user", "password")
    if "mssql" not in st.secrets:
        raise ValueError(".streamlit/secrets.toml에 [mssql] 설정이 없습니다.")
    config = dict(st.secrets["mssql"])
    if not clean_text(config.get("driver")):
        config["driver"] = config.get("dirver", "")
    missing = [key for key in required + ("driver",) if not clean_text(config.get(key))]
    if missing:
        raise ValueError(f"[mssql] 필수 항목이 없습니다: {', '.join(missing)}")
    return config


def fetch_metadata_mappings(config: dict[str, Any]) -> pd.DataFrame:
    mapping_owner = oracle_identifier(clean_text(config["owner"]))
    pool = get_oracle_pool(
        clean_text(config["host"]),
        int(config["port"]),
        clean_text(config["service_name"]),
        clean_text(config["user"]),
        clean_text(config["password"]),
        2,
    )
    sql = f"""
        SELECT DISTINCT
               TRIM(SRC_SYSTEM) AS SRC_SYSTEM,
               TRIM(SRC_TABLE) AS SRC_TABLE,
               TRIM(SRC_ENTITY) AS SRC_ENTITY,
               TRIM(TGT_TABLE) AS TGT_TABLE
          FROM {mapping_owner}.{oracle_identifier(MAPPING_TABLE)}
         WHERE UPPER(TRIM(MIG_YN)) = :mig_yn
         ORDER BY TRIM(SRC_SYSTEM), TRIM(SRC_TABLE), TRIM(TGT_TABLE)
    """
    with pool.acquire() as connection:
        with connection.cursor() as cursor:
            cursor.execute(sql, mig_yn="Y")
            rows = cursor.fetchall()
            columns = [item[0] for item in cursor.description]
    return normalize_mappings(pd.DataFrame(rows, columns=columns))


def normalize_mappings(raw: pd.DataFrame) -> pd.DataFrame:
    required = {"SRC_SYSTEM", "SRC_TABLE", "SRC_ENTITY", "TGT_TABLE"}
    renamed = raw.copy()
    renamed.columns = [str(column).replace("\ufeff", "").strip().upper() for column in raw.columns]
    missing = sorted(required - set(renamed.columns))
    if missing:
        raise ValueError(f"입력 파일에 필요한 컬럼이 없습니다: {', '.join(missing)}")
    result = renamed[["SRC_SYSTEM", "SRC_TABLE", "SRC_ENTITY", "TGT_TABLE"]].copy()
    for column in result.columns:
        result[column] = result[column].map(upper_text)
    result = result[(result["SRC_SYSTEM"] != "") & (result["SRC_TABLE"] != "") & (result["TGT_TABLE"] != "")]
    result = result.drop_duplicates().sort_values(["SRC_SYSTEM", "SRC_TABLE", "TGT_TABLE"]).reset_index(drop=True)
    return result


def normalize_excel_mappings(raw: pd.DataFrame) -> pd.DataFrame:
    renamed = raw.copy()
    renamed.columns = [str(column).replace("\ufeff", "").strip().upper() for column in raw.columns]
    required = {"SYSTEM", "TABLE"}
    missing = sorted(required - set(renamed.columns))
    if missing:
        raise ValueError(f"엑셀 입력 컬럼이 없습니다: {', '.join(missing)}")
    result = pd.DataFrame(
        {
            "SRC_SYSTEM": renamed["SYSTEM"].map(upper_text),
            "SRC_TABLE": renamed["TABLE"].map(upper_text),
        }
    )
    result = result[(result["SRC_SYSTEM"] != "") & (result["SRC_TABLE"] != "")].copy()
    result["SRC_ENTITY"] = ""
    result["TGT_TABLE"] = result.apply(lambda row: derived_target_table(row["SRC_SYSTEM"], row["SRC_TABLE"]), axis=1)
    return result.drop_duplicates().sort_values(["SRC_SYSTEM", "SRC_TABLE", "TGT_TABLE"]).reset_index(drop=True)


def load_excel_mappings(uploaded_file: Any) -> pd.DataFrame:
    if uploaded_file.name.lower().endswith(".csv"):
        raw = pd.read_csv(uploaded_file, dtype=str)
    else:
        raw = pd.read_excel(uploaded_file, dtype=str)
    return normalize_excel_mappings(raw)


def as_mapping_rows(frame: pd.DataFrame) -> list[MappingRow]:
    return [
        MappingRow(
            source_db=upper_text(row.SRC_SYSTEM),
            source_table=upper_text(row.SRC_TABLE),
            source_entity=upper_text(row.SRC_ENTITY),
            target_table=upper_text(row.TGT_TABLE),
        )
        for row in frame.itertuples(index=False)
    ]


def mssql_connection(config: dict[str, Any], database: str) -> Any:
    if pyodbc is None:
        raise RuntimeError("pyodbc가 설치되지 않았습니다. requirements.txt 설치 후 다시 실행하십시오.")
    connection_string = (
        f"DRIVER={{{clean_text(config['driver'])}}};"
        f"SERVER={clean_text(config['host'])},{clean_text(config['port'])};"
        f"DATABASE={database};UID={clean_text(config['user'])};PWD={clean_text(config['password'])};"
        "TrustServerCertificate=yes;"
    )
    return pyodbc.connect(connection_string, timeout=10)


def fetch_mssql_measurement(mapping: MappingRow, config: dict[str, Any]) -> dict[str, Any]:
    source_table = mssql_identifier(mapping.source_table)
    object_name = source_table
    with mssql_connection(config, mapping.source_db) as connection:
        cursor = connection.cursor()
        cursor.execute(f"SELECT COUNT_BIG(1) FROM {source_table}")
        count_row = cursor.fetchone()
        cursor.execute(
            """
            SELECT
                COALESCE(SUM(CASE WHEN index_id IN (0, 1)
                    THEN in_row_data_page_count + lob_used_page_count + row_overflow_used_page_count
                    ELSE 0 END), 0) AS data_pages,
                COALESCE(SUM(used_page_count), 0) AS used_pages
              FROM sys.dm_db_partition_stats
             WHERE object_id = OBJECT_ID(?)
            """,
            object_name,
        )
        size_row = cursor.fetchone()
        cursor.execute(
            """
            SELECT TOP (1)
                   COALESCE(CONVERT(nvarchar(4000), property_value.value), table_info.name)
              FROM sys.tables AS table_info
              LEFT JOIN sys.extended_properties AS property_value
                ON property_value.class = 1
               AND property_value.major_id = table_info.object_id
               AND property_value.minor_id = 0
               AND property_value.name = 'MS_Description'
             WHERE table_info.object_id = OBJECT_ID(?)
            """,
            object_name,
        )
        entity_row = cursor.fetchone()
    data_mb = float(size_row[0] or 0) * 8 / 1024
    used_mb = float(size_row[1] or 0) * 8 / 1024
    return {
        "count": int(count_row[0]),
        "data_mb": data_mb,
        "index_mb": max(used_mb - data_mb, 0.0),
        "total_mb": used_mb,
        "entity": clean_text(entity_row[0]) if entity_row else "",
        "error": "",
    }


def fetch_oracle_count(mapping: MappingRow, pool: oracledb.ConnectionPool) -> dict[str, Any]:
    qualified_table = f"{oracle_identifier(TARGET_OWNER)}.{oracle_identifier(target_table_name(mapping.target_table))}"
    with pool.acquire() as connection:
        with connection.cursor() as cursor:
            cursor.execute(f"SELECT COUNT(1) FROM {qualified_table}")
            row = cursor.fetchone()
    return {"count": int(row[0]), "error": ""}


def run_mssql_task(mapping: MappingRow, config: dict[str, Any]) -> tuple[str, MappingRow, dict[str, Any]]:
    try:
        return "mssql", mapping, fetch_mssql_measurement(mapping, config)
    except Exception as exc:
        return "mssql", mapping, {"count": None, "data_mb": None, "index_mb": None, "total_mb": None, "error": str(exc)}


def run_oracle_task(mapping: MappingRow, pool: oracledb.ConnectionPool) -> tuple[str, MappingRow, dict[str, Any]]:
    try:
        return "oracle", mapping, fetch_oracle_count(mapping, pool)
    except Exception as exc:
        return "oracle", mapping, {"count": None, "error": str(exc)}


def mismatch_count(source_values: dict[MappingRow, dict[str, Any]], target_values: dict[MappingRow, dict[str, Any]]) -> int:
    return sum(
        1
        for mapping, source in source_values.items()
        if mapping in target_values
        and source.get("count") is not None
        and target_values[mapping].get("count") is not None
        and source["count"] != target_values[mapping]["count"]
    )


def status_text(
    name: str,
    completed: int,
    total: int,
    started: str,
    finished: str,
    errors: int,
    elapsed: str,
    include_date: bool = False,
) -> str:
    percent = (completed / total * 100) if total else 100.0
    if errors:
        state, state_class = "오류", "ops-state-error"
    elif completed < total:
        state, state_class = "진행중", "ops-state-running"
    else:
        state, state_class = "완료", "ops-state-done"
    error_class = "ops-error-active" if errors else "ops-error-clear"
    name_text = f'<span class="ops-progress-name">{escape(name)}</span> ' if name else ""
    started_label = clean_text(started) if include_date else time_label(started)
    finished_label = clean_text(finished) if include_date else time_label(finished)
    level_class = "ops-progress-summary" if include_date else "ops-progress-detail"
    return (
        f'<div class="{level_class}">{name_text}<span class="ops-state {state_class}">{state}</span> '
        f'<span class="ops-progress-count">· {completed:,}/{total:,} ({percent:.1f}%)</span> '
        f'<span class="ops-progress-time">· {escape(started_label)} ~ {escape(finished_label)} [{escape(elapsed)}]</span> '
        f'<span class="{error_class}">오류({errors:,})</span></div>'
    )


def comparison_status_text(completed: int, total: int, mismatches: int, errors: int) -> str:
    if errors or mismatches:
        badge, icon = ":red-badge[검증 실패]", ":material/error:"
    elif completed < total:
        badge, icon = ":orange-badge[검증 진행중]", ":material/hourglass_top:"
    else:
        badge, icon = ":green-badge[검증 완료]", ":material/check_circle:"
    mismatch_text = f":red[불일치 {mismatches:,}]" if mismatches else ":gray[불일치 0]"
    error_text = f":red[오류 {errors:,}]" if errors else ":gray[오류 0]"
    return f"{icon} **검증 상태** {badge} · 양쪽 완료 {completed:,}/{total:,} · {mismatch_text} · {error_text}"


def render_progress_dashboard(progress_state: dict[str, Any] | None) -> None:
    if not progress_state:
        return
    mssql_state = progress_state["mssql"]
    oracle_state = progress_state["oracle"]
    with st.container(border=True, gap=None):
        render_section_title("▥", "조회 진행률")
        render_fixed_gap(20)
        mssql_panel, oracle_panel = st.columns(2, gap="medium")
        with mssql_panel:
            st.markdown('<div class="ops-system-title">▣ MSSQL</div>', unsafe_allow_html=True)
            render_fixed_gap(10)
            st.markdown(
                status_text(
                    "",
                    mssql_state["completed"],
                    mssql_state["total"],
                    mssql_state["started"],
                    mssql_state["finished"],
                    mssql_state["errors"],
                    mssql_state["elapsed"],
                    include_date=True,
                ),
                unsafe_allow_html=True,
            )
            render_fixed_gap(10)
            st.progress(mssql_state["completed"] / mssql_state["total"] if mssql_state["total"] else 1.0)
            render_fixed_gap(10)
            detail_rows = "".join(
                status_text(
                    database,
                    values["completed"],
                    values["total"],
                    values["started"],
                    values["finished"],
                    values["errors"],
                    values["elapsed"],
                )
                for database, values in progress_state["databases"].items()
            )
            st.markdown(f'<div class="ops-progress-list">{detail_rows}</div>', unsafe_allow_html=True)
            render_fixed_gap(14)
        with oracle_panel:
            st.markdown('<div class="ops-system-title">▣ ORACLE</div>', unsafe_allow_html=True)
            render_fixed_gap(10)
            st.markdown(
                status_text(
                    "",
                    oracle_state["completed"],
                    oracle_state["total"],
                    oracle_state["started"],
                    oracle_state["finished"],
                    oracle_state["errors"],
                    oracle_state["elapsed"],
                    include_date=True,
                ),
                unsafe_allow_html=True,
            )
            render_fixed_gap(10)
            st.progress(oracle_state["completed"] / oracle_state["total"] if oracle_state["total"] else 1.0)
            render_fixed_gap(14)


def run_comparison(
    mappings: list[MappingRow],
    oracle: dict[str, Any],
    mssql: dict[str, Any],
    oracle_workers: int,
    snapshot_slot: Any,
    progress_slot: Any,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    run_started = now_label()
    pool = get_oracle_pool(
        clean_text(oracle["host"]),
        int(oracle["port"]),
        clean_text(oracle["service_name"]),
        clean_text(oracle["user"]),
        clean_text(oracle["password"]),
        max(oracle_workers, 1),
    )
    source_started = run_started
    target_started = run_started
    running_summary = {
        "mssql": {"started": source_started, "finished": "-", "elapsed": "-", "mismatches": 0, "errors": 0},
        "oracle": {"started": target_started, "finished": "-", "elapsed": "-", "mismatches": 0, "errors": 0},
    }
    with snapshot_slot.container():
        render_operations_snapshot(running_summary, len(mappings))
    database_totals = {database: 0 for database in sorted({item.source_db for item in mappings})}
    for mapping in mappings:
        database_totals[mapping.source_db] = database_totals.get(mapping.source_db, 0) + 1
    database_progress = {
        database: {"completed": 0, "total": total, "started": source_started, "finished": "-", "errors": 0, "elapsed": "00:00:00"}
        for database, total in database_totals.items()
    }
    progress_state = {
        "mssql": {"completed": 0, "total": len(mappings), "started": source_started, "finished": "-", "errors": 0, "elapsed": "00:00:00"},
        "oracle": {"completed": 0, "total": len(mappings), "started": target_started, "finished": "-", "errors": 0, "elapsed": "00:00:00"},
        "databases": database_progress,
    }
    st.session_state.countcheck_progress = progress_state
    with progress_slot.container():
        render_progress_dashboard(progress_state)

    source_values: dict[MappingRow, dict[str, Any]] = {}
    target_values: dict[MappingRow, dict[str, Any]] = {}
    oracle_completed = 0
    source_completed = 0
    source_finished = "-"
    target_finished = "-"
    futures: dict[Future[tuple[str, MappingRow, dict[str, Any]]], str] = {}
    with ExitStack() as stack:
        source_executors = {
            database: stack.enter_context(ThreadPoolExecutor(max_workers=1, thread_name_prefix=f"mssql-{database.lower()}"))
            for database in database_totals
        }
        target_executor = stack.enter_context(ThreadPoolExecutor(max_workers=oracle_workers, thread_name_prefix="oracle-count"))
        for mapping in mappings:
            futures[source_executors[mapping.source_db].submit(run_mssql_task, mapping, mssql)] = "mssql"
            futures[target_executor.submit(run_oracle_task, mapping, pool)] = "oracle"
        pending = set(futures)
        while pending:
            done, pending = wait(pending, return_when=FIRST_COMPLETED)
            for future in done:
                side, mapping, result = future.result()
                if side == "mssql":
                    source_values[mapping] = result
                    source_completed += 1
                    database_progress[mapping.source_db]["completed"] += 1
                    if clean_text(result.get("error")):
                        database_progress[mapping.source_db]["errors"] += 1
                    if database_progress[mapping.source_db]["completed"] == database_progress[mapping.source_db]["total"]:
                        database_progress[mapping.source_db]["finished"] = now_label()
                        database_progress[mapping.source_db]["elapsed"] = elapsed_label(
                            database_progress[mapping.source_db]["started"], database_progress[mapping.source_db]["finished"]
                        )
                    if source_completed == len(mappings) and source_finished == "-":
                        source_finished = now_label()
                else:
                    target_values[mapping] = result
                    oracle_completed += 1
                    if oracle_completed == len(mappings) and target_finished == "-":
                        target_finished = now_label()
            current_mismatches = mismatch_count(source_values, target_values)
            source_errors = sum(bool(clean_text(result.get("error"))) for result in source_values.values())
            target_errors = sum(bool(clean_text(result.get("error"))) for result in target_values.values())
            running_summary = {
                "mssql": {
                    "started": source_started,
                    "finished": source_finished,
                    "elapsed": elapsed_label(source_started, source_finished),
                    "mismatches": current_mismatches,
                    "errors": source_errors,
                },
                "oracle": {
                    "started": target_started,
                    "finished": target_finished,
                    "elapsed": elapsed_label(target_started, target_finished),
                    "mismatches": current_mismatches,
                    "errors": target_errors,
                },
            }
            with snapshot_slot.container():
                render_operations_snapshot(running_summary, len(mappings))
            progress_state = {
                "mssql": {
                    "completed": source_completed,
                    "total": len(mappings),
                    "started": source_started,
                    "finished": source_finished,
                    "errors": source_errors,
                    "elapsed": elapsed_label(source_started, source_finished),
                },
                "oracle": {
                    "completed": oracle_completed,
                    "total": len(mappings),
                    "started": target_started,
                    "finished": target_finished,
                    "errors": target_errors,
                    "elapsed": elapsed_label(target_started, target_finished),
                },
                "databases": database_progress,
            }
            st.session_state.countcheck_progress = progress_state
            with progress_slot.container():
                render_progress_dashboard(progress_state)

    rows: list[dict[str, Any]] = []
    for mapping in mappings:
        source = source_values.get(mapping, {})
        target = target_values.get(mapping, {})
        source_error = clean_text(source.get("error"))
        target_error = clean_text(target.get("error"))
        source_count = source.get("count")
        target_count = target.get("count")
        difference = source_count - target_count if source_count is not None and target_count is not None else None
        note = " | ".join(part for part in (source_error, target_error) if part)
        status = "성공" if difference == 0 and not note else "실패"
        rows.append(
            {
                "소스DB": mapping.source_db,
                "소스TABLE": mapping.source_table,
                "소스ENTITY": source.get("entity") or mapping.source_entity,
                "타겟TABLE": mapping.target_table,
                "검증결과": status,
                "오류여부": "오류" if note else "정상",
                "소스CNT": source_count,
                "타겟CNT": target_count,
                "차이": difference,
                "소스DATA MB": source.get("data_mb"),
                "소스DATA GB": (source.get("data_mb") / 1024) if source.get("data_mb") is not None else None,
                "소스INDEX MB": source.get("index_mb"),
                "소스합계 MB": source.get("total_mb"),
                "비고": note,
            }
        )
    result_frame = pd.DataFrame(rows, columns=RESULT_COLUMNS)
    mismatch_count_value = int(result_frame["차이"].notna().mul(result_frame["차이"].ne(0)).sum())
    summary = {
        "mssql": {
            "started": source_started,
            "finished": source_finished,
            "elapsed": elapsed_label(source_started, source_finished),
            "processed": source_completed,
            "mismatches": mismatch_count_value,
            "errors": int(sum(bool(clean_text(result.get("error"))) for result in source_values.values())),
        },
        "oracle": {
            "started": target_started,
            "finished": target_finished,
            "elapsed": elapsed_label(target_started, target_finished),
            "processed": oracle_completed,
            "mismatches": mismatch_count_value,
            "errors": int(sum(bool(clean_text(result.get("error"))) for result in target_values.values())),
        },
    }
    return result_frame, summary


def render_summary(summary: dict[str, Any] | None) -> None:
    if not summary:
        return
    frame = pd.DataFrame(
        [
            {"구분": "MSSQL", **summary["mssql"]},
            {"구분": "ORACLE", **summary["oracle"]},
        ]
    ).rename(
        columns={
            "started": "조회시작",
            "finished": "조회종료",
            "elapsed": "소요시간",
            "processed": "진행건수",
            "mismatches": "불일치건수",
            "errors": "오류건수",
        }
    )
    st.dataframe(
        frame,
        hide_index=True,
        width="stretch",
        column_config={
            "진행건수": st.column_config.NumberColumn(format="localized"),
            "불일치건수": st.column_config.NumberColumn(format="localized"),
            "오류건수": st.column_config.NumberColumn(format="localized"),
        },
    )


def result_cell_style(value: object) -> str:
    if clean_text(value) == "성공":
        return "color: #69d68a; font-weight: 700;"
    if clean_text(value) == "실패":
        return "color: #ff7c86; font-weight: 700;"
    return ""


def error_cell_style(value: object) -> str:
    if clean_text(value) == "오류":
        return "color: #ff7c86; font-weight: 700;"
    if clean_text(value) == "정상":
        return "color: #97aabd;"
    return ""


def result_excel_bytes(frame: pd.DataFrame) -> bytes:
    from openpyxl.styles import Border, Font, PatternFill, Side

    export_frame = frame.copy()
    for column in ("소스DB", "소스TABLE", "소스ENTITY", "타겟TABLE"):
        if column in export_frame:
            export_frame[column] = export_frame[column].map(lambda value: "" if pd.isna(value) else str(value))
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_frame.to_excel(writer, sheet_name="검증결과", index=False)
        worksheet = writer.sheets["검증결과"]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        thin_side = Side(style="thin", color="9EADBF")
        all_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        body_font = Font(name="맑은 고딕", size=10)
        header_font = Font(name="맑은 고딕", size=10, bold=True)
        header_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
        for row_index, row in enumerate(worksheet.iter_rows(), start=1):
            for cell in row:
                cell.font = header_font if row_index == 1 else body_font
                cell.border = all_border
                if row_index == 1:
                    cell.fill = header_fill
        for column_cells in worksheet.columns:
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 42)
        header_columns = {cell.value: cell.column for cell in worksheet[1]}
        for header in ("소스DB", "소스TABLE", "소스ENTITY", "타겟TABLE"):
            if header in header_columns:
                for row in worksheet.iter_rows(min_row=2, min_col=header_columns[header], max_col=header_columns[header]):
                    row[0].number_format = "@"
        for header in ("소스CNT", "타겟CNT", "차이"):
            if header in header_columns:
                for row in worksheet.iter_rows(min_row=2, min_col=header_columns[header], max_col=header_columns[header]):
                    row[0].number_format = "#,##0"
        for header in ("소스DATA MB", "소스DATA GB", "소스INDEX MB", "소스합계 MB"):
            if header in header_columns:
                for row in worksheet.iter_rows(min_row=2, min_col=header_columns[header], max_col=header_columns[header]):
                    row[0].number_format = "#,##0.00000"
    return output.getvalue()


def render_results() -> None:
    results = st.session_state.countcheck_results
    result_section = st.container(border=True, gap=None)
    render_section_title("▣", "검증결과", result_section)
    if results.empty:
        result_section.caption("조회 결과가 없습니다.")
        return
    render_fixed_gap(20, result_section)
    filter_database, filter_table, filter_entity, filter_result, filter_error = result_section.columns(5)
    database_options = ["전체", *sorted(results["소스DB"].dropna().unique())]
    status_options = ["전체", "성공", "실패"]
    error_options = ["전체", "정상", "오류"]
    selected_database = filter_database.selectbox("소스DB", database_options, key="countcheck_filter_db")
    table_keyword = filter_table.text_input("소스TABLE", key="countcheck_filter_table")
    entity_keyword = filter_entity.text_input("소스ENTITY", key="countcheck_filter_entity")
    selected_status = filter_result.selectbox("검증결과", status_options, key="countcheck_filter_status")
    selected_error = filter_error.selectbox("오류여부", error_options, key="countcheck_filter_error")
    filtered = results.copy()
    if selected_database != "전체":
        filtered = filtered[filtered["소스DB"] == selected_database]
    if table_keyword.strip():
        keyword = table_keyword.strip().upper()
        filtered = filtered[filtered["소스TABLE"].astype(str).str.contains(keyword, case=False, na=False)]
    if entity_keyword.strip():
        keyword = entity_keyword.strip().upper()
        filtered = filtered[filtered["소스ENTITY"].astype(str).str.contains(keyword, case=False, na=False)]
    if selected_status != "전체":
        filtered = filtered[filtered["검증결과"] == selected_status]
    if selected_error != "전체":
        filtered = filtered[filtered["오류여부"] == selected_error]
    styled_results = filtered.style.map(result_cell_style, subset=["검증결과"]).map(error_cell_style, subset=["오류여부"])
    result_section.download_button(
        "조회 결과 엑셀 다운로드",
        data=result_excel_bytes(filtered),
        file_name=f"MSSQL_ORACLE_검증결과_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        icon=":material/download:",
        width="content",
    )
    render_fixed_gap(20, result_section)
    result_section.dataframe(
        styled_results,
        hide_index=True,
        width="stretch",
        column_config={
            "검증결과": st.column_config.TextColumn(pinned=True),
            "소스CNT": st.column_config.NumberColumn(format="localized"),
            "타겟CNT": st.column_config.NumberColumn(format="localized"),
            "차이": st.column_config.NumberColumn(format="localized"),
            "소스DATA MB": st.column_config.NumberColumn(format="localized"),
            "소스DATA GB": st.column_config.NumberColumn(format="localized"),
            "소스INDEX MB": st.column_config.NumberColumn(format="localized"),
            "소스합계 MB": st.column_config.NumberColumn(format="localized"),
        },
    )


def main() -> None:
    st.set_page_config(page_title="MSSQL vs Oracle Count Check", page_icon=":material/fact_check:", layout="wide")
    init_state()
    apply_operations_style()
    main_content = st.container(gap=None)
    with main_content:
        render_operations_header()
        snapshot_slot = st.empty()
        snapshot_target_gap_slot = st.empty()
        target_slot = st.empty()
        target_progress_gap_slot = st.empty()
        progress_slot = st.empty()
        progress_results_gap_slot = st.empty()
        results_slot = st.empty()

    with st.sidebar:
        st.header(":material/tune: 조회 대상")
        source_mode = st.segmented_control("목록 원본", ["메타데이터", "엑셀"], default="메타데이터")
        uploaded_file = None
        if source_mode == "엑셀":
            uploaded_file = st.file_uploader("매핑 파일", type=["xlsx", "xls", "csv"])
        oracle_workers = st.selectbox("Oracle 워커 수", options=list(range(1, 21)), index=4)
        load_clicked = st.button("대상 목록 조회", icon=":material/playlist_add_check:", width="stretch")
        run_clicked = st.button("검증 시작", icon=":material/play_arrow:", type="primary", width="stretch")

    if load_clicked:
        try:
            if source_mode == "엑셀":
                if uploaded_file is None:
                    raise ValueError("엑셀 또는 CSV 파일을 선택하십시오.")
                mappings = load_excel_mappings(uploaded_file)
            else:
                mappings = fetch_metadata_mappings(oracle_config())
            st.session_state.countcheck_mappings = mappings
            st.session_state.countcheck_results = empty_result_frame()
            st.session_state.countcheck_summary = None
            st.session_state.countcheck_progress = None
            st.toast(f"{len(mappings):,}건의 대상 목록을 불러왔습니다.", icon=":material/check_circle:")
        except Exception as exc:
            st.error(str(exc), icon=":material/error:")

    mappings = st.session_state.countcheck_mappings
    with snapshot_slot.container():
        render_operations_snapshot(st.session_state.countcheck_summary, len(mappings))
    if not mappings.empty:
        render_fixed_gap(38, snapshot_target_gap_slot)
        with target_slot.container():
            with st.container(border=True, gap=None):
                render_section_title("▦", "검증 대상")
                render_fixed_gap(20)
                database_counts = mappings.groupby("SRC_SYSTEM").size().sort_index()
                target_cards = tuple((f"▣ {database}", f"{table_count:,}건") for database, table_count in database_counts.items())
                render_card_row(target_cards)
                render_fixed_gap(20)
        render_fixed_gap(38, target_progress_gap_slot)
    else:
        snapshot_target_gap_slot.empty()
        target_slot.empty()
        target_progress_gap_slot.empty()

    if st.session_state.countcheck_progress:
        with progress_slot.container():
            render_progress_dashboard(st.session_state.countcheck_progress)
    else:
        progress_slot.empty()

    if run_clicked:
        try:
            if mappings.empty:
                raise ValueError("먼저 검증 대상 목록을 조회하십시오.")
            invalid_databases = sorted(set(mappings["SRC_SYSTEM"]) - set(SOURCE_DATABASES))
            if invalid_databases:
                raise ValueError(f"지원 대상 MSSQL DB가 아닙니다: {', '.join(invalid_databases)}")
            result_frame, summary = run_comparison(
                as_mapping_rows(mappings),
                oracle_config(),
                mssql_config(),
                oracle_workers,
                snapshot_slot,
                progress_slot,
            )
            st.session_state.countcheck_results = result_frame
            st.session_state.countcheck_summary = summary
            with snapshot_slot.container():
                render_operations_snapshot(summary, len(mappings))
            st.toast("검증이 완료되었습니다.", icon=":material/check_circle:")
        except Exception as exc:
            st.error(str(exc), icon=":material/error:")

    if st.session_state.countcheck_summary is not None or not st.session_state.countcheck_results.empty:
        render_fixed_gap(38, progress_results_gap_slot)
        with results_slot.container():
            render_results()
    else:
        progress_results_gap_slot.empty()
        results_slot.empty()


if __name__ == "__main__":
    main()
