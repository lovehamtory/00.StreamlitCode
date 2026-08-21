from __future__ import annotations

import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

import pandas as pd
import streamlit as st

try:
    import psycopg
except ImportError:
    psycopg = None


MAPPING_TABLE = "TB_MIG_TABLE_INFO"
MAPPING_COLUMNS = ["SRC_SYSTEM", "SRC_TABLE", "SRC_ENTITY", "TGT_TABLE", "SM_CHK_YN", "HSH_CHK_YN"]
RESULT_COLUMNS = ["소스SYSTEM", "소스TABLE", "소스ENTITY", "타겟TABLE", "SUM 검증", "HASH 검증", "검증결과", "오류여부", "소스CNT", "타겟CNT", "CNT 차이", "SUM 대상 컬럼", "SUM 불일치", "HASH 대상 컬럼", "소스 HASH 합계", "타겟 HASH 합계", "비고"]
DETAIL_COLUMNS = ["소스SYSTEM", "소스TABLE", "타겟TABLE", "검증항목", "컬럼", "소스값", "타겟값", "차이", "검증결과", "비고"]
HASH_MISMATCH_COLUMNS = ["소스SYSTEM", "소스TABLE", "타겟TABLE", "구분", "PK 컬럼", "PK 값", "소스 행 해시", "타겟 행 해시", "비고"]
NUMERIC_TYPES = {"smallint", "integer", "bigint", "decimal", "numeric", "real", "double precision", "float4", "float8"}
TEXT_TYPES = {"character varying", "character", "varchar", "char", "text", "nchar", "nvarchar"}
DATE_TYPES = {"date", "timestamp without time zone", "timestamp with time zone", "timestamp", "timestamptz"}


@dataclass(frozen=True)
class MappingRow:
    system: str
    source_table: str
    entity: str
    target_table: str
    sum_yn: str
    hash_yn: str


def clean(value: object) -> str:
    return "" if value is None or pd.isna(value) else str(value).strip()


def upper(value: object) -> str:
    return re.sub(r"\s+", " ", clean(value)).upper()


def enabled(value: object) -> bool:
    return upper(value) in {"Y", "YES", "1", "TRUE"}


def identifier(value: str) -> str:
    name = clean(value)
    if not name or "\x00" in name or len(name.encode("utf-8")) > 127:
        raise ValueError(f"식별자 형식이 올바르지 않습니다: {value}")
    return '"' + name.replace('"', '""') + '"'


def split_table(value: str, default_schema: str) -> tuple[str, str]:
    parts = [part.strip().strip('"') for part in clean(value).split(".") if part.strip()]
    if len(parts) == 1:
        return default_schema, parts[0]
    if len(parts) == 2:
        return parts[0], parts[1]
    raise ValueError(f"스키마.테이블 형식이 올바르지 않습니다: {value}")


def qualified(schema_name: str, table_name: str) -> str:
    return f"{identifier(schema_name)}.{identifier(table_name)}"


def now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def elapsed(started: str, finished: str) -> str:
    try:
        end = datetime.strptime(finished, "%Y-%m-%d %H:%M:%S") if finished != "-" else datetime.now()
        seconds = max(0, int((end - datetime.strptime(started, "%Y-%m-%d %H:%M:%S")).total_seconds()))
        hours, seconds = divmod(seconds, 3600)
        minutes, seconds = divmod(seconds, 60)
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    except ValueError:
        return "-"


def init_state() -> None:
    for key, value in {
        "gp_rs_mappings": pd.DataFrame(columns=MAPPING_COLUMNS),
        "gp_rs_results": pd.DataFrame(columns=RESULT_COLUMNS),
        "gp_rs_details": pd.DataFrame(columns=DETAIL_COLUMNS),
        "gp_rs_hash_mismatches": pd.DataFrame(columns=HASH_MISMATCH_COLUMNS),
        "gp_rs_summary": None,
        "gp_rs_progress": None,
    }.items():
        st.session_state.setdefault(key, value)


def apply_green_style() -> None:
    st.markdown("""<style>
    [data-testid="stAppViewContainer"]{background:radial-gradient(circle at 18% -10%,#1d6648 0,#10271d 35%,#07140f 100%)}
    [data-testid="stHeader"]{background:transparent}.block-container{max-width:1520px;padding-top:2.1rem}
    .hero{padding:.8rem 1.05rem;margin-bottom:.8rem;background:linear-gradient(105deg,#195b40,#102b20);border:1px solid #34815d;border-left:4px solid #7de8ac;border-radius:10px;color:#f7fffa}.hero h1{margin:0;font-size:1.28rem;line-height:1.2;letter-spacing:-.02em}.hero p{margin:.3rem 0 0;color:#b9dcc8;font-size:.72rem}.kicker{color:#9de8bb;font-size:.68rem;font-weight:700;letter-spacing:.1em}
    .card{height:78px;padding:12px;background:#173224;border:1px solid #3e7555;border-radius:9px;color:#f7fffa}.card b{display:block;color:#c5e8d1;font-size:13px}.card span{display:block;margin-top:12px;font-weight:750}
    [data-testid="stMetric"]{background:#173224;border:1px solid #3e7555;border-radius:9px}.stButton>button{border-color:#4a9a6d;background:#1f6848;color:#fff}.stButton>button[kind="primary"]{background:linear-gradient(100deg,#267c55,#4fba7a)}
    [data-testid="stProgressBar"]>div>div{background:linear-gradient(90deg,#2f8d61,#83dc9f)}</style>""", unsafe_allow_html=True)


def config(section: str) -> dict[str, Any]:
    required = ("host", "port", "database", "user", "password")
    if section not in st.secrets:
        raise ValueError(f".streamlit/secrets.toml에 [{section}] 설정이 없습니다.")
    values = dict(st.secrets[section])
    missing = [key for key in required if not clean(values.get(key))]
    if missing:
        raise ValueError(f"[{section}] 필수 항목이 없습니다: {', '.join(missing)}")
    return values


def connect(values: dict[str, Any]) -> Any:
    if psycopg is None:
        raise RuntimeError(f"psycopg가 현재 실행 Python에 설치되지 않았습니다: {sys.executable}")
    arguments: dict[str, Any] = {"host": clean(values["host"]), "port": int(values["port"]), "dbname": clean(values["database"]), "user": clean(values["user"]), "password": clean(values["password"]), "connect_timeout": int(values.get("connect_timeout", 15))}
    if clean(values.get("sslmode")):
        arguments["sslmode"] = clean(values["sslmode"])
    return psycopg.connect(**arguments)


def metadata_settings(target: dict[str, Any]) -> tuple[str, str]:
    settings = dict(st.secrets.get("migration_metadata", {}))
    return clean(settings.get("schema")) or clean(target.get("default_schema")) or "public", clean(settings.get("table")) or MAPPING_TABLE


def normalize_mappings(frame: pd.DataFrame) -> pd.DataFrame:
    copied = frame.copy()
    copied.columns = [str(column).replace("\ufeff", "").strip().upper() for column in copied.columns]
    missing = sorted(set(MAPPING_COLUMNS) - set(copied.columns))
    if missing:
        raise ValueError(f"입력 데이터에 필요한 컬럼이 없습니다: {', '.join(missing)}")
    result = copied[MAPPING_COLUMNS].copy()
    for column in result.columns:
        result[column] = result[column].map(upper)
    return result[(result.SRC_SYSTEM != "") & (result.SRC_TABLE != "") & (result.TGT_TABLE != "")].drop_duplicates().sort_values(["SRC_SYSTEM", "SRC_TABLE", "TGT_TABLE"]).reset_index(drop=True)


def fetch_metadata(target: dict[str, Any]) -> pd.DataFrame:
    schema_name, table_name = metadata_settings(target)
    query = f"SELECT DISTINCT TRIM(SRC_SYSTEM), TRIM(SRC_TABLE), TRIM(SRC_ENTITY), TRIM(TGT_TABLE), COALESCE(TRIM(SM_CHK_YN),'N'), COALESCE(TRIM(HSH_CHK_YN),'N') FROM {qualified(schema_name, table_name)} WHERE UPPER(TRIM(MIG_YN))=%s ORDER BY TRIM(SRC_SYSTEM), TRIM(SRC_TABLE), TRIM(TGT_TABLE)"
    with connect(target) as connection:
        with connection.cursor() as cursor:
            cursor.execute(query, ("Y",))
            return normalize_mappings(pd.DataFrame(cursor.fetchall(), columns=MAPPING_COLUMNS))


def load_excel(uploaded: Any) -> pd.DataFrame:
    raw = pd.read_csv(uploaded, dtype=str) if uploaded.name.lower().endswith(".csv") else pd.read_excel(uploaded, dtype=str)
    return normalize_mappings(raw)


def rows(frame: pd.DataFrame) -> list[MappingRow]:
    return [MappingRow(upper(row.SRC_SYSTEM), upper(row.SRC_TABLE), clean(row.SRC_ENTITY), upper(row.TGT_TABLE), upper(row.SM_CHK_YN), upper(row.HSH_CHK_YN)) for row in frame.itertuples(index=False)]


def columns(values: dict[str, Any], schema_name: str, table_name: str) -> dict[str, tuple[str, str]]:
    with connect(values) as connection:
        with connection.cursor() as cursor:
            cursor.execute("SELECT column_name, data_type FROM information_schema.columns WHERE table_schema=%s AND table_name=%s ORDER BY ordinal_position", (schema_name, table_name))
            return {clean(name).lower():(clean(name), clean(data_type).lower()) for name, data_type in cursor.fetchall()}


def comparable(mapping: MappingRow, source: dict[str, Any], target: dict[str, Any]) -> tuple[tuple[str, str, str, str], dict[str, list[str]]]:
    source_schema, source_table = split_table(mapping.source_table, clean(source.get("default_schema")) or "public")
    target_schema, target_table = split_table(mapping.target_table, clean(target.get("default_schema")) or "public")
    source_columns, target_columns = columns(source, source_schema, source_table), columns(target, target_schema, target_table)
    if not source_columns or not target_columns:
        raise ValueError("원천 또는 대상 테이블을 찾지 못했거나 조회 권한이 없습니다.")
    common = [key for key in source_columns if key in target_columns]
    return (source_schema, source_table, target_schema, target_table), {
        "SUM": [source_columns[key][0] for key in common if source_columns[key][1] in NUMERIC_TYPES and target_columns[key][1] in NUMERIC_TYPES] if enabled(mapping.sum_yn) else [],
        "HASH": [source_columns[key][0] for key in common] if enabled(mapping.hash_yn) else [],
    }


def metrics(values: dict[str, Any], schema_name: str, table_name: str, groups: dict[str, list[str]]) -> dict[tuple[str, str], Any]:
    expressions, keys = ["COUNT(*)"], [("COUNT", "")]
    for column in groups["SUM"]:
        expressions.append(f"SUM({identifier(column)})"); keys.append(("SUM", column))
    with connect(values) as connection:
        with connection.cursor() as cursor:
            cursor.execute(f"SELECT {', '.join(expressions)} FROM {qualified(schema_name, table_name)}")
            result = dict(zip(keys, cursor.fetchone()))
            if groups["HASH"]:
                encoded_columns = [f"COALESCE(LENGTH(CAST({identifier(column)} AS VARCHAR(65535)))::VARCHAR || ':' || CAST({identifier(column)} AS VARCHAR(65535)), '-1:')" for column in groups["HASH"]]
                cursor.execute(f"SELECT MD5({' || CHR(30) || '.join(encoded_columns)}) FROM {qualified(schema_name, table_name)}")
                total = 0
                while batch := cursor.fetchmany(10000):
                    total += sum(int(row[0], 16) for row in batch)
                result[("HASH SUM", "")] = total
            return result


def primary_key_columns(values: dict[str, Any], schema_name: str, table_name: str) -> list[str]:
    query = """
        SELECT kcu.column_name
          FROM information_schema.table_constraints tc
          JOIN information_schema.key_column_usage kcu
            ON kcu.constraint_name = tc.constraint_name
           AND kcu.constraint_schema = tc.constraint_schema
           AND kcu.table_name = tc.table_name
         WHERE tc.constraint_type = 'PRIMARY KEY'
           AND tc.table_schema = %s
           AND tc.table_name = %s
         ORDER BY kcu.ordinal_position
    """
    with connect(values) as connection:
        with connection.cursor() as cursor:
            cursor.execute(query, (schema_name, table_name))
            return [text(row[0]) for row in cursor.fetchall()]


def hash_expression(hash_columns: list[str]) -> str:
    encoded = [f"COALESCE(LENGTH(CAST({identifier(column)} AS VARCHAR(65535)))::VARCHAR || ':' || CAST({identifier(column)} AS VARCHAR(65535)), '-1:')" for column in hash_columns]
    return f"MD5({' || CHR(30) || '.join(encoded)})"


def hash_rows(values: dict[str, Any], schema_name: str, table_name: str, pk_columns: list[str], hash_columns: list[str]) -> dict[tuple[str, ...], str]:
    selected = ", ".join(identifier(column) for column in pk_columns)
    query = f"SELECT {selected}, {hash_expression(hash_columns)} FROM {qualified(schema_name, table_name)}"
    result: dict[tuple[str, ...], str] = {}
    with connect(values) as connection:
        with connection.cursor() as cursor:
            cursor.execute(query)
            while batch := cursor.fetchmany(10000):
                for row in batch:
                    key = tuple("<NULL>" if value is None else str(value) for value in row[:-1])
                    result[key] = str(row[-1])
    return result


def hash_mismatch_rows(values_by_mapping: dict[MappingRow, dict[str, Any]], mappings: list[MappingRow], source: dict[str, Any], target: dict[str, Any], maximum: int = 1000) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for mapping in mappings:
        values = values_by_mapping.get(mapping, {})
        if not enabled(mapping.hash_yn) or values.get("error") or not values.get("groups", {}).get("HASH"):
            continue
        if equal(values["source"].get(("HASH SUM", "")), values["target"].get(("HASH SUM", ""))):
            continue
        source_schema, source_table, target_schema, target_table = values["locations"]
        try:
            source_pk = primary_key_columns(source, source_schema, source_table)
            target_pk = primary_key_columns(target, target_schema, target_table)
            if [normalized(item) for item in source_pk] != [normalized(item) for item in target_pk] or not source_pk:
                rows.append({"소스SYSTEM": mapping.system, "소스TABLE": mapping.source_table, "타겟TABLE": mapping.target_table, "구분": "PK 확인 실패", "PK 컬럼": ", ".join(source_pk) or "-", "PK 값": "-", "소스 행 해시": "", "타겟 행 해시": "", "비고": "양쪽 PK가 없거나 PK 컬럼 구성이 다릅니다."})
                continue
            source_hashes = hash_rows(source, source_schema, source_table, source_pk, values["groups"]["HASH"])
            target_hashes = hash_rows(target, target_schema, target_table, source_pk, values["groups"]["HASH"])
            for key in sorted(set(source_hashes) | set(target_hashes)):
                source_hash, target_hash = source_hashes.get(key), target_hashes.get(key)
                if source_hash == target_hash:
                    continue
                status = "원천만 존재" if target_hash is None else "대상만 존재" if source_hash is None else "값 불일치"
                rows.append({"소스SYSTEM": mapping.system, "소스TABLE": mapping.source_table, "타겟TABLE": mapping.target_table, "구분": status, "PK 컬럼": ", ".join(source_pk), "PK 값": " | ".join(key), "소스 행 해시": source_hash or "", "타겟 행 해시": target_hash or "", "비고": ""})
                if len(rows) >= maximum:
                    return pd.DataFrame(rows, columns=HASH_MISMATCH_COLUMNS)
        except Exception as error:
            rows.append({"소스SYSTEM": mapping.system, "소스TABLE": mapping.source_table, "타겟TABLE": mapping.target_table, "구분": "PK 추출 오류", "PK 컬럼": "-", "PK 값": "-", "소스 행 해시": "", "타겟 행 해시": "", "비고": str(error)})
    return pd.DataFrame(rows, columns=HASH_MISMATCH_COLUMNS)


def equal(left: Any, right: Any) -> bool:
    if left is None or right is None:
        return left is right
    try:
        return Decimal(str(left)) == Decimal(str(right))
    except (InvalidOperation, ValueError):
        return str(left) == str(right)


def difference(left: Any, right: Any) -> Any:
    if left is None or right is None:
        return None
    try:
        return Decimal(str(left)) - Decimal(str(right))
    except (InvalidOperation, ValueError):
        return "" if equal(left, right) else "상이"


def validate(mapping: MappingRow, source: dict[str, Any], target: dict[str, Any]) -> tuple[MappingRow, dict[str, Any]]:
    result: dict[str, Any] = {"source": {}, "target": {}, "groups": {"SUM": [], "HASH": []}, "locations": (), "error": ""}
    try:
        locations, result["groups"] = comparable(mapping, source, target)
        result["locations"] = locations
        result["source"] = metrics(source, locations[0], locations[1], result["groups"])
        result["target"] = metrics(target, locations[2], locations[3], result["groups"])
        for flag, group in ((mapping.sum_yn, "SUM"), (mapping.hash_yn, "HASH")):
            if enabled(flag) and not result["groups"][group]:
                result["error"] = " | ".join(filter(None, [result["error"], f"{group} 검증 대상 공통 컬럼이 없습니다."]))
    except Exception as error:
        result["error"] = str(error)
    return mapping, result


def group_status(group: str, enabled_flag: bool, values: dict[str, Any]) -> tuple[str, int]:
    if not enabled_flag: return "미사용", 0
    if values["error"]: return "오류", 0
    keys = [(group, column) for column in values["groups"][group]]
    if not keys: return "대상없음", 0
    mismatch = sum(not equal(values["source"].get(key), values["target"].get(key)) for key in keys)
    return ("성공" if mismatch == 0 else "불일치"), mismatch


def create_frames(values_by_mapping: dict[MappingRow, dict[str, Any]], mappings: list[MappingRow]) -> tuple[pd.DataFrame, pd.DataFrame]:
    results, details = [], []
    for mapping in mappings:
        values = values_by_mapping.get(mapping, {"source": {}, "target": {}, "groups": {"SUM": [], "HASH": []}, "error": "결과가 없습니다."})
        count_ok = not values["error"] and equal(values["source"].get(("COUNT", "")), values["target"].get(("COUNT", "")))
        sum_status, sum_bad = group_status("SUM", enabled(mapping.sum_yn), values)
        hash_status, _ = group_status("HASH", enabled(mapping.hash_yn), values)
        if enabled(mapping.hash_yn) and not values["error"] and values["groups"]["HASH"]:
            hash_status = "성공" if equal(values["source"].get(("HASH SUM", "")), values["target"].get(("HASH SUM", ""))) else "불일치"
        success = count_ok and sum_status in {"성공", "미사용"} and hash_status in {"성공", "미사용"}
        results.append({"소스SYSTEM":mapping.system,"소스TABLE":mapping.source_table,"소스ENTITY":mapping.entity,"타겟TABLE":mapping.target_table,"SUM 검증":sum_status,"HASH 검증":hash_status,"검증결과":"성공" if success else "실패","오류여부":"오류" if values["error"] else "정상","소스CNT":values["source"].get(("COUNT", "")),"타겟CNT":values["target"].get(("COUNT", "")),"CNT 차이":difference(values["source"].get(("COUNT", "")),values["target"].get(("COUNT", ""))),"SUM 대상 컬럼":", ".join(values["groups"]["SUM"]),"SUM 불일치":sum_bad,"HASH 대상 컬럼":", ".join(values["groups"]["HASH"]),"소스 HASH 합계":values["source"].get(("HASH SUM", "")),"타겟 HASH 합계":values["target"].get(("HASH SUM", "")),"비고":values["error"]})
        detail_keys = [("COUNT", ""), *[("SUM", column) for column in values["groups"]["SUM"]]]
        if enabled(mapping.hash_yn):
            detail_keys.append(("HASH SUM", ""))
        for group, column in detail_keys:
            left, right = values["source"].get((group,column)), values["target"].get((group,column))
            details.append({"소스SYSTEM":mapping.system,"소스TABLE":mapping.source_table,"타겟TABLE":mapping.target_table,"검증항목":group,"컬럼":column or "-","소스값":left,"타겟값":right,"차이":difference(left,right),"검증결과":"오류" if values["error"] else ("성공" if equal(left,right) else "불일치"),"비고":values["error"]})
    return pd.DataFrame(results, columns=RESULT_COLUMNS), pd.DataFrame(details, columns=DETAIL_COLUMNS)


def run_comparison(mappings: list[MappingRow], source: dict[str, Any], target: dict[str, Any], workers: int, snapshot: Any, progress_slot: Any) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    started = now(); values_by_mapping: dict[MappingRow, dict[str, Any]] = {}; completed = 0
    state = {name:{"started":started,"finished":"-","elapsed":"00:00:00","completed":0,"total":len(mappings),"errors":0,"mismatches":0} for name in ("greenplum","redshift")}
    with ThreadPoolExecutor(max_workers=workers, thread_name_prefix="gp-rs-check") as executor:
        futures = [executor.submit(validate, mapping, source, target) for mapping in mappings]
        for future in as_completed(futures):
            mapping, values = future.result(); values_by_mapping[mapping] = values; completed += 1
            for name in state:
                state[name]["completed"] = completed; state[name]["errors"] = sum(bool(item["error"]) for item in values_by_mapping.values()); state[name]["elapsed"] = elapsed(started,"-")
                if completed == len(mappings): state[name]["finished"] = now(); state[name]["elapsed"] = elapsed(started,state[name]["finished"])
            current, _ = create_frames(values_by_mapping, mappings); mismatch = int(current["검증결과"].eq("실패").sum()) if not current.empty else 0
            for name in state: state[name]["mismatches"] = mismatch
            with snapshot.container(): render_snapshot(state, len(mappings))
            st.session_state.gp_rs_progress = state
            with progress_slot.container(): render_progress(state)
    result, details = create_frames(values_by_mapping, mappings)
    mismatches = hash_mismatch_rows(values_by_mapping, mappings, source, target)
    return result, details, mismatches, state


def render_snapshot(summary: dict[str, Any] | None, total: int) -> None:
    source, target = (summary or {}).get("greenplum",{}), (summary or {}).get("redshift",{})
    cards = (("▣ 대상 테이블",f"{total:,}건"),("◷ 조회 시작",source.get("started","-")),("◷ 조회 종료",target.get("finished","-")),("◴ 소요 시간",target.get("elapsed","-")),("⇄ 불일치",f"{int(target.get('mismatches',0)):,}건"),("⚠ 오류",f"{int(target.get('errors',0)):,}건"))
    for column, card in zip(st.columns(6,gap="small"),cards):
        with column: st.markdown(f'<div class="card"><b>{card[0]}</b><span>{card[1]}</span></div>',unsafe_allow_html=True)


def render_progress(state: dict[str, Any] | None) -> None:
    if not state: return
    with st.container(border=True):
        st.markdown("#### ▥ 조회 진행률")
        for column, name in zip(st.columns(2), ("greenplum","redshift")):
            item=state[name]
            with column:
                st.markdown(f"**▣ {name.upper()}** · {item['completed']:,}/{item['total']:,} · 오류 {item['errors']:,}")
                st.progress(item["completed"]/item["total"] if item["total"] else 1.0)
                st.caption(f"{item['started']} ~ {item['finished']} [{item['elapsed']}]")


def excel_bytes(results: pd.DataFrame, details: pd.DataFrame) -> bytes:
    from openpyxl.styles import Border, Font, PatternFill, Side
    output=BytesIO()
    with pd.ExcelWriter(output,engine="openpyxl") as writer:
        for frame,name in ((results,"검증결과"),(details,"검증상세")):
            frame.to_excel(writer,sheet_name=name,index=False); sheet=writer.sheets[name]; sheet.freeze_panes="A2"; sheet.auto_filter.ref=sheet.dimensions
            for row_number,row in enumerate(sheet.iter_rows(),start=1):
                for cell in row: cell.font=Font(name="맑은 고딕",size=10,bold=row_number==1); cell.border=Border(left=Side(style="thin",color="B29BC8"),right=Side(style="thin",color="B29BC8"),top=Side(style="thin",color="B29BC8"),bottom=Side(style="thin",color="B29BC8")); cell.fill=PatternFill(fill_type="solid",fgColor="D9C2F0") if row_number==1 else PatternFill()
            for cells in sheet.columns: sheet.column_dimensions[cells[0].column_letter].width=min(max(len(str(cell.value or "")) for cell in cells)+2,42)
    return output.getvalue()


def render_results() -> None:
    results,details,hash_mismatches=st.session_state.gp_rs_results,st.session_state.gp_rs_details,st.session_state.gp_rs_hash_mismatches
    with st.container(border=True):
        st.markdown("#### ▣ 검증결과")
        system,table,status,error=st.columns(4); selected_system=system.selectbox("소스SYSTEM",["전체",*sorted(results["소스SYSTEM"].unique())],key="gp_system"); keyword=table.text_input("소스TABLE",key="gp_table"); selected_status=status.selectbox("검증결과",["전체","성공","실패"],key="gp_status"); selected_error=error.selectbox("오류여부",["전체","정상","오류"],key="gp_error")
        filtered=results.copy()
        if selected_system!="전체": filtered=filtered[filtered["소스SYSTEM"]==selected_system]
        if keyword: filtered=filtered[filtered["소스TABLE"].str.contains(keyword,case=False,na=False)]
        if selected_status!="전체": filtered=filtered[filtered["검증결과"]==selected_status]
        if selected_error!="전체": filtered=filtered[filtered["오류여부"]==selected_error]
        st.download_button("조회 결과 엑셀 다운로드",data=excel_bytes(filtered,details),file_name=f"GREENPLUM_REDSHIFT_검증결과_{datetime.now():%Y%m%d_%H%M%S}.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",icon=":material/download:")
        st.dataframe(filtered,hide_index=True,column_config={column:st.column_config.NumberColumn(format="localized") for column in ("소스CNT","타겟CNT","CNT 차이","SUM 불일치","소스 HASH 합계","타겟 HASH 합계")})
        with st.expander("컬럼별 검증 상세"):
            st.dataframe(details,hide_index=True)
        if not hash_mismatches.empty:
            st.markdown("#### :material/key: 해시 불일치 PK")
            st.dataframe(hash_mismatches, hide_index=True)


def main() -> None:
    st.set_page_config(page_title="Migration 결과 검증",page_icon=":material/fact_check:",layout="wide")
    init_state(); apply_green_style(); source,target=config("greenplum"),config("redshift_sql")
    st.markdown('<div class="hero"><h1>✦ Migration 결과 검증</h1><p>⚙️ Created by ♡홍율파파♡</p></div>',unsafe_allow_html=True)
    snapshot,target_slot,progress_slot,result_slot=st.empty(),st.empty(),st.empty(),st.empty()
    with st.sidebar:
        st.header(":material/tune: 조회 대상"); mode=st.segmented_control("목록 원본",["메타데이터","엑셀"],default="메타데이터"); uploaded=st.file_uploader("매핑 파일",type=["xlsx","xls","csv"]) if mode=="엑셀" else None; workers=st.selectbox("검증 워커 수",list(range(1,21)),index=4); load_clicked=st.button("대상 목록 조회",icon=":material/playlist_add_check:",width="stretch"); run_clicked=st.button("검증 시작",icon=":material/play_arrow:",type="primary",width="stretch")
    if load_clicked:
        try:
            if mode=="엑셀" and uploaded is None: raise ValueError("엑셀 또는 CSV 파일을 선택하십시오.")
            mappings=load_excel(uploaded) if mode=="엑셀" else fetch_metadata(target); st.session_state.gp_rs_mappings=mappings; st.session_state.gp_rs_results=pd.DataFrame(columns=RESULT_COLUMNS); st.session_state.gp_rs_details=pd.DataFrame(columns=DETAIL_COLUMNS); st.session_state.gp_rs_hash_mismatches=pd.DataFrame(columns=HASH_MISMATCH_COLUMNS); st.session_state.gp_rs_summary=None; st.session_state.gp_rs_progress=None; st.toast(f"{len(mappings):,}건의 대상 목록을 불러왔습니다.",icon=":material/check_circle:")
        except Exception as error: st.error(str(error),icon=":material/error:")
    mappings=st.session_state.gp_rs_mappings
    with snapshot.container(): render_snapshot(st.session_state.gp_rs_summary,len(mappings))
    if not mappings.empty:
        with target_slot.container():
            with st.container(border=True):
                st.markdown("#### ▦ 검증 대상")
                for column,(system,count) in zip(st.columns(6,gap="small"),mappings.groupby("SRC_SYSTEM").size().sort_index().items()):
                    with column: st.markdown(f'<div class="card"><b>▣ {system}</b><span>{count:,}건</span></div>',unsafe_allow_html=True)
    if st.session_state.gp_rs_progress:
        with progress_slot.container(): render_progress(st.session_state.gp_rs_progress)
    if run_clicked:
        try:
            if mappings.empty: raise ValueError("먼저 검증 대상 목록을 조회하십시오.")
            results,details,hash_mismatches,summary=run_comparison(rows(mappings),source,target,workers,snapshot,progress_slot); st.session_state.gp_rs_results=results; st.session_state.gp_rs_details=details; st.session_state.gp_rs_hash_mismatches=hash_mismatches; st.session_state.gp_rs_summary=summary; st.toast("검증이 완료되었습니다.",icon=":material/check_circle:")
        except Exception as error: st.error(str(error),icon=":material/error:")
    if st.session_state.gp_rs_summary is not None or not st.session_state.gp_rs_results.empty:
        with result_slot.container(): render_results()


if __name__ == "__main__":
    main()
