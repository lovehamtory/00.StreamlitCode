from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any, Callable

import pandas as pd
import streamlit as st


ARTIFACT_ROOT = Path(__file__).parent.parent / "artifact"
LAYOUTS: dict[str, dict[str, Any]] = {
    "TBL_DFN": {"name": "테이블정의서", "sheet": "테이블정의서", "items": [("MPG_ID", "테이블매핑ID"), ("PRJ_CD", "프로젝트코드"), ("SBJ_AREA_CD", "주제영역코드"), ("TGT_CONN_ID", "대상접속ID"), ("TGT_SCH_NM", "대상스키마명"), ("TGT_TBL_NM", "대상테이블명"), ("TGT_TBL_CMT", "대상테이블설명"), ("SRC_CONN_ID", "원천접속ID"), ("SRC_SCH_NM", "원천스키마명"), ("SRC_TBL_NM", "원천테이블명"), ("LOAD_STS_CD", "기본적재상태"), ("SYS_COL_NM_ARR", "시스템컬럼명"), ("SYS_COL_FMT_CD", "시스템컬럼형식"), ("INCR_MTHD_CD", "증분방식"), ("SRC_INCR_COL_NM_ARR", "원천증분컬럼명"), ("PARL_MTHD_CD", "S3병렬방식"), ("PARL_CND_ARR", "S3병렬조건"), ("META_VER_NO", "메타버전번호")]},
    "COL_DFN": {"name": "컬럼정의서", "sheet": "컬럼정의서", "items": [("MPG_ID", "테이블매핑ID"), ("PRJ_CD", "프로젝트코드"), ("SBJ_AREA_CD", "주제영역코드"), ("TGT_SCH_NM", "대상스키마명"), ("TGT_TBL_NM", "대상테이블명"), ("SRC_SCH_NM", "원천스키마명"), ("SRC_TBL_NM", "원천테이블명"), ("COL_ORD", "매핑순서"), ("TGT_COL_NO", "대상컬럼순번"), ("TGT_COL_NM", "대상컬럼명"), ("TGT_COL_CMT", "대상컬럼설명"), ("TGT_DATA_TYPE", "대상데이터타입"), ("TGT_NULL_YN", "대상NULL허용"), ("TGT_KEY_ROLE_CD", "대상키역할"), ("COL_MPG_MTHD_CD", "컬럼매핑방식"), ("TGT_EXPR", "이행적용SQL식"), ("DFLT_EXPR", "이행기본값SQL식"), ("S3_COL_NM", "S3중간컬럼명"), ("S3_DATA_TYPE", "S3중간데이터타입"), ("SRC_EXPR", "이관적용SQL식"), ("SRC_REF_COL_NM_ARR", "원천참조컬럼명"), ("SRC_COL_NO", "원천컬럼순번"), ("SRC_COL_NM", "원천컬럼명"), ("SRC_DATA_TYPE", "원천데이터타입"), ("SRC_NULL_YN", "원천NULL허용"), ("SRC_KEY_ROLE_CD", "원천키역할"), ("SUM_VALD_YN", "SUM검증여부"), ("HSH_VALD_YN", "HASH검증여부")]},
    "MPG_DFN": {"name": "매핑정의서", "sheet": "매핑정의서", "items": [("MPG_ID", "테이블매핑ID"), ("PRJ_CD", "프로젝트코드"), ("SBJ_AREA_CD", "주제영역코드"), ("TGT_CONN_ID", "대상접속ID"), ("TGT_SCH_NM", "대상스키마명"), ("TGT_TBL_NM", "대상테이블명"), ("SRC_CONN_ID", "원천접속ID"), ("SRC_SCH_NM", "원천스키마명"), ("SRC_TBL_NM", "원천테이블명"), ("COL_ORD", "매핑순서"), ("TGT_COL_NM", "대상컬럼명"), ("TGT_DATA_TYPE", "대상데이터타입"), ("COL_MPG_MTHD_CD", "컬럼매핑방식"), ("TGT_EXPR", "이행적용SQL식"), ("DFLT_EXPR", "이행기본값SQL식"), ("S3_COL_NM", "S3중간컬럼명"), ("SRC_EXPR", "이관적용SQL식"), ("SRC_REF_COL_NM_ARR", "원천참조컬럼명"), ("SRC_COL_NM", "원천컬럼명"), ("SUM_VALD_YN", "SUM검증여부"), ("HSH_VALD_YN", "HASH검증여부")]},
    "UTEST_RSLT": {"name": "단위테스트결과서", "sheet": "테이블별로그", "items": [("DAG_NM", "DAG명"), ("DAG_RUN_ID", "DAG실행ID"), ("MPG_ID", "테이블매핑ID"), ("SRC_TBL", "원천테이블"), ("TGT_TBL", "대상테이블"), ("TASK_NM", "태스크명"), ("WRK_DVSN_CD", "작업구분"), ("LOAD_MTHD_CD", "적재방식"), ("WRK_STS_CD", "작업상태"), ("SRC_ROW_CNT", "원천건수"), ("TGT_ROW_CNT", "대상건수"), ("S3_BYTE_SIZE", "S3크기"), ("S3_MNF_PATH", "S3매니페스트"), ("SQL_FILE_PATH", "SQL경로"), ("SRC_WHERE_CND", "원천조회조건"), ("WRK_STT_DTM", "작업시작일시"), ("WRK_END_DTM", "작업종료일시"), ("WRK_ELPS_SEC", "실행경과초"), ("WRK_MSG", "작업메시지")]},
    "ITEST_RSLT": {"name": "통합테스트결과서", "sheet": "DAG별로그", "items": [("DAG_NM", "DAG명"), ("DAG_RUN_ID", "DAG실행ID"), ("DAG_DVSN_CD", "DAG구분"), ("MAP_CNT", "전체테이블건수"), ("SUC_CNT", "완료테이블건수"), ("RUN_CNT", "진행테이블건수"), ("ERR_CNT", "오류테이블건수"), ("WRK_STS_CD", "작업상태"), ("WRK_STT_DTM", "작업시작일시"), ("WRK_END_DTM", "작업종료일시"), ("WRK_ELPS_SEC", "실행경과초"), ("WRK_MSG", "작업메시지")]},
    "VALD_RSLT": {"name": "검증결과서", "sheet": "검증결과", "items": [("DAG_NM", "DAG명"), ("DAG_RUN_ID", "DAG실행ID"), ("MPG_ID", "테이블매핑ID"), ("VALD_DVSN_CD", "검증구분"), ("S3_MANF_PATH", "S3매니페스트"), ("CNT_VALD_STS_CD", "건수검증"), ("SRC_CNT", "원천건수"), ("TGT_CNT", "대상건수"), ("CNT_DIFF", "건수차이"), ("SUM_VALD_STS_CD", "SUM검증"), ("HSH_VALD_STS_CD", "HASH검증"), ("VALD_STS_CD", "검증상태"), ("VALD_STT_DTM", "검증시작일시"), ("VALD_END_DTM", "검증종료일시"), ("VALD_ELPS_SEC", "검증경과초"), ("VALD_MSG", "검증메시지")]},
}


def excel_bytes(sheets: list[tuple[str, pd.DataFrame]]) -> bytes:
    from openpyxl.styles import Border, Font, PatternFill, Side

    output = BytesIO()
    border = Border(left=Side(style="thin", color="D7DEE8"), right=Side(style="thin", color="D7DEE8"), top=Side(style="thin", color="D7DEE8"), bottom=Side(style="thin", color="D7DEE8"))
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, frame in sheets:
            frame.to_excel(writer, sheet_name=sheet_name[:31], index=False)
            sheet = writer.sheets[sheet_name[:31]]
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for row_number, row in enumerate(sheet.iter_rows(), start=1):
                for cell in row:
                    cell.font = Font(name="맑은 고딕", size=10, bold=row_number == 1)
                    cell.border = border
                    if row_number == 1:
                        cell.fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
            for cells in sheet.columns:
                sheet.column_dimensions[cells[0].column_letter].width = min(max(max(len(str(cell.value or "")) for cell in cells) + 2, 11), 42)
    return output.getvalue()


def read_layout(query_frame: Callable[..., pd.DataFrame], values: dict[str, Any], schema_name: str, qualified: Callable[[str, str], str], layout_code: str) -> pd.DataFrame:
    return query_frame(values, f'''SELECT item_id AS "ITEM_ID", item_nm AS "ITEM_NM", disp_ord AS "DISP_ORD", out_yn AS "OUT_YN"
                                   FROM {qualified(schema_name, 'tb_mig_artf_item')}
                                  WHERE artf_dvsn_cd = %s
                                  ORDER BY disp_ord, item_id''', (layout_code,))


def configured_items(configured: pd.DataFrame, layout_code: str) -> pd.DataFrame:
    result = pd.DataFrame(LAYOUTS[layout_code]["items"], columns=["ITEM_ID", "ITEM_NM"])
    result["DISP_ORD"] = range(1, len(result) + 1)
    result["OUT_YN"] = True
    if configured.empty:
        return result
    current = configured.set_index("ITEM_ID")
    for index, row in result.iterrows():
        if row.ITEM_ID in current.index:
            result.at[index, "ITEM_NM"] = current.at[row.ITEM_ID, "ITEM_NM"]
            result.at[index, "DISP_ORD"] = int(current.at[row.ITEM_ID, "DISP_ORD"])
            result.at[index, "OUT_YN"] = bool(current.at[row.ITEM_ID, "OUT_YN"])
    return result


def save_layout(connect: Callable[..., Any], values: dict[str, Any], schema_name: str, qualified: Callable[[str, str], str], layout_code: str, items: pd.DataFrame) -> None:
    if items.ITEM_NM.astype(str).str.strip().eq("").any() or items.DISP_ORD.duplicated().any() or items.DISP_ORD.lt(1).any():
        raise ValueError("산출물 항목명과 출력순서를 확인하십시오.")
    with connect(values) as connection:
        with connection.cursor() as cursor:
            cursor.execute(f"DELETE FROM {qualified(schema_name, 'tb_mig_artf_item')} WHERE artf_dvsn_cd = %s", (layout_code,))
            for row in items.sort_values("DISP_ORD").itertuples(index=False):
                cursor.execute(f"INSERT INTO {qualified(schema_name, 'tb_mig_artf_item')} (artf_dvsn_cd, item_id, item_nm, disp_ord, out_yn) VALUES (%s, %s, %s, %s, %s)", (layout_code, row.ITEM_ID, row.ITEM_NM, int(row.DISP_ORD), bool(row.OUT_YN)))
        connection.commit()


def mapping_frame(query_frame: Callable[..., pd.DataFrame], values: dict[str, Any], schema_name: str, qualified: Callable[[str, str], str], columns: bool = False) -> pd.DataFrame:
    if not columns:
        return query_frame(values, f'''SELECT T.mpg_id AS "MPG_ID", T.prj_cd AS "PRJ_CD", T.sbj_area_cd AS "SBJ_AREA_CD", A.src_conn_id AS "SRC_CONN_ID", T.src_sch_nm AS "SRC_SCH_NM", T.src_tbl_nm AS "SRC_TBL_NM", A.tgt_conn_id AS "TGT_CONN_ID", T.tgt_sch_nm AS "TGT_SCH_NM", T.tgt_tbl_nm AS "TGT_TBL_NM", T.tgt_tbl_cmt AS "TGT_TBL_CMT", T.load_sts_cd AS "LOAD_STS_CD", T.sys_col_nm_arr AS "SYS_COL_NM_ARR", T.sys_col_fmt_cd AS "SYS_COL_FMT_CD", T.incr_mthd_cd AS "INCR_MTHD_CD", T.src_incr_col_nm_arr AS "SRC_INCR_COL_NM_ARR", T.parl_mthd_cd AS "PARL_MTHD_CD", T.parl_cnd_arr AS "PARL_CND_ARR", T.meta_ver_no AS "META_VER_NO"
                                         FROM {qualified(schema_name, 'tb_mig_tbl_mpg')} T JOIN {qualified(schema_name, 'tb_mig_sbj_area')} A ON A.sbj_area_cd = T.sbj_area_cd
                                        WHERE T.active_yn = TRUE
                                        ORDER BY T.prj_cd, T.sbj_area_cd, T.mpg_id''')
    return query_frame(values, f'''SELECT T.mpg_id AS "MPG_ID", T.prj_cd AS "PRJ_CD", T.sbj_area_cd AS "SBJ_AREA_CD", T.tgt_sch_nm AS "TGT_SCH_NM", T.tgt_tbl_nm AS "TGT_TBL_NM", T.src_sch_nm AS "SRC_SCH_NM", T.src_tbl_nm AS "SRC_TBL_NM", C.col_ord AS "COL_ORD", C.tgt_col_no AS "TGT_COL_NO", C.tgt_col_nm AS "TGT_COL_NM", C.tgt_col_cmt AS "TGT_COL_CMT", C.tgt_data_type AS "TGT_DATA_TYPE", C.tgt_null_yn AS "TGT_NULL_YN", C.tgt_key_role_cd AS "TGT_KEY_ROLE_CD", C.col_mpg_mthd_cd AS "COL_MPG_MTHD_CD", C.tgt_expr AS "TGT_EXPR", C.dflt_expr AS "DFLT_EXPR", C.s3_col_nm AS "S3_COL_NM", C.s3_data_type AS "S3_DATA_TYPE", C.src_expr AS "SRC_EXPR", C.src_ref_col_nm_arr AS "SRC_REF_COL_NM_ARR", C.src_col_no AS "SRC_COL_NO", C.src_col_nm AS "SRC_COL_NM", C.src_data_type AS "SRC_DATA_TYPE", C.src_null_yn AS "SRC_NULL_YN", C.src_key_role_cd AS "SRC_KEY_ROLE_CD", C.sum_vald_yn AS "SUM_VALD_YN", C.hsh_vald_yn AS "HSH_VALD_YN"
                                   FROM {qualified(schema_name, 'tb_mig_col_mpg')} C
                                   JOIN {qualified(schema_name, 'tb_mig_tbl_mpg')} T ON T.mpg_id = C.mpg_id
                                  WHERE T.active_yn = TRUE AND C.active_yn = TRUE
                                  ORDER BY T.prj_cd, T.sbj_area_cd, T.mpg_id, C.col_ord''')


def runtime_frame(document_code: str, query_frame: Callable[..., pd.DataFrame], values: dict[str, Any], schema_name: str, qualified: Callable[[str, str], str], start_dt: date, end_dt: date) -> pd.DataFrame:
    if document_code == "UTEST_RSLT":
        return query_frame(values, f'''SELECT L.dag_nm AS "DAG_NM", L.dag_run_id AS "DAG_RUN_ID", L.mpg_id AS "MPG_ID", COALESCE(T.src_sch_nm || '.' || T.src_tbl_nm, '') AS "SRC_TBL", COALESCE(T.tgt_sch_nm || '.' || T.tgt_tbl_nm, '') AS "TGT_TBL", L.task_nm AS "TASK_NM", L.wrk_dvsn_cd AS "WRK_DVSN_CD", L.load_mthd_cd AS "LOAD_MTHD_CD", L.wrk_sts_cd AS "WRK_STS_CD", L.src_row_cnt AS "SRC_ROW_CNT", L.tgt_row_cnt AS "TGT_ROW_CNT", L.s3_byte_size AS "S3_BYTE_SIZE", L.s3_mnf_path AS "S3_MNF_PATH", L.sql_file_path AS "SQL_FILE_PATH", L.src_where_cnd AS "SRC_WHERE_CND", L.wrk_stt_dtm AS "WRK_STT_DTM", L.wrk_end_dtm AS "WRK_END_DTM", L.wrk_elps_sec AS "WRK_ELPS_SEC", L.wrk_msg AS "WRK_MSG"
                                         FROM {qualified(schema_name, 'tb_mig_run_log')} L LEFT JOIN {qualified(schema_name, 'tb_mig_tbl_mpg')} T ON T.mpg_id = L.mpg_id
                                        WHERE CAST(L.wrk_stt_dtm AS DATE) BETWEEN %s AND %s ORDER BY L.run_hist_id DESC''', (start_dt, end_dt))
    if document_code == "ITEST_RSLT":
        return query_frame(values, f'''SELECT dag_nm AS "DAG_NM", dag_run_id AS "DAG_RUN_ID", dag_dvsn_cd AS "DAG_DVSN_CD", map_cnt AS "MAP_CNT", suc_cnt AS "SUC_CNT", run_cnt AS "RUN_CNT", err_cnt AS "ERR_CNT", wrk_sts_cd AS "WRK_STS_CD", wrk_stt_dtm AS "WRK_STT_DTM", wrk_end_dtm AS "WRK_END_DTM", wrk_elps_sec AS "WRK_ELPS_SEC", wrk_msg AS "WRK_MSG"
                                         FROM {qualified(schema_name, 'vw_mig_dag_run_sum')}
                                        WHERE CAST(wrk_stt_dtm AS DATE) BETWEEN %s AND %s ORDER BY dag_exec_id DESC''', (start_dt, end_dt))
    return query_frame(values, f'''SELECT dag_nm AS "DAG_NM", dag_run_id AS "DAG_RUN_ID", mpg_id AS "MPG_ID", vald_dvsn_cd AS "VALD_DVSN_CD", s3_manf_path AS "S3_MANF_PATH", cnt_vald_sts_cd AS "CNT_VALD_STS_CD", src_cnt AS "SRC_CNT", tgt_cnt AS "TGT_CNT", cnt_diff AS "CNT_DIFF", sum_vald_sts_cd AS "SUM_VALD_STS_CD", hsh_vald_sts_cd AS "HSH_VALD_STS_CD", vald_sts_cd AS "VALD_STS_CD", vald_stt_dtm AS "VALD_STT_DTM", vald_end_dtm AS "VALD_END_DTM", vald_elps_sec AS "VALD_ELPS_SEC", vald_msg AS "VALD_MSG"
                                   FROM {qualified(schema_name, 'tb_mig_vald_rslt')}
                                  WHERE CAST(vald_stt_dtm AS DATE) BETWEEN %s AND %s ORDER BY vald_hist_id DESC''', (start_dt, end_dt))


def apply_layout(frame: pd.DataFrame, items: pd.DataFrame) -> pd.DataFrame:
    selected = items.loc[items.OUT_YN].sort_values("DISP_ORD")
    if frame.empty:
        return pd.DataFrame(columns=selected.ITEM_NM.tolist())
    source = {str(column).upper(): column for column in frame.columns}
    pairs = [(row.ITEM_ID, row.ITEM_NM) for row in selected.itertuples(index=False) if row.ITEM_ID in source]
    output = frame.loc[:, [source[code] for code, _ in pairs]].copy()
    output.columns = [name for _, name in pairs]
    return output.where(pd.notna(output), None)


def render_artifacts(values: dict[str, Any], schema_name: str, query_frame: Callable[..., pd.DataFrame], connect: Callable[..., Any], qualified: Callable[[str, str], str]) -> None:
    mode = st.segmented_control("업무", ["산출물 생성", "레이아웃 정의"], default="산출물 생성", label_visibility="collapsed")
    if mode == "레이아웃 정의":
        code = st.selectbox("산출물", list(LAYOUTS), format_func=lambda item: LAYOUTS[item]["name"])
        try:
            edited = st.data_editor(configured_items(read_layout(query_frame, values, schema_name, qualified, code), code), hide_index=True, disabled=["ITEM_ID"], column_config={"ITEM_ID": "항목코드", "ITEM_NM": "항목명", "DISP_ORD": st.column_config.NumberColumn("출력순서", min_value=1), "OUT_YN": st.column_config.CheckboxColumn("출력")})
            if st.button("저장", type="primary", icon=":material/save:"):
                save_layout(connect, values, schema_name, qualified, code, edited)
                st.success("레이아웃을 저장했습니다.", icon=":material/check_circle:")
        except Exception as error:
            st.error(f"레이아웃 처리 실패: {error}", icon=":material/error:")
        return
    left, right, action = st.columns([2, 1, 1])
    with left:
        code = st.selectbox("산출물", list(LAYOUTS), format_func=lambda item: LAYOUTS[item]["name"])
    with right:
        start_dt = st.date_input("시작일", value=date.today() - timedelta(days=7))
        end_dt = st.date_input("종료일", value=date.today())
    with action:
        requested = st.button("생성", type="primary", icon=":material/description:")
    if requested:
        try:
            if end_dt < start_dt:
                raise ValueError("종료일은 시작일보다 빠를 수 없습니다.")
            if code == "TBL_DFN":
                frame = mapping_frame(query_frame, values, schema_name, qualified)
            elif code in {"COL_DFN", "MPG_DFN"}:
                frame = mapping_frame(query_frame, values, schema_name, qualified, columns=True)
            else:
                frame = runtime_frame(code, query_frame, values, schema_name, qualified, start_dt, end_dt)
            items = configured_items(read_layout(query_frame, values, schema_name, qualified, code), code)
            data = excel_bytes([(LAYOUTS[code]["sheet"], apply_layout(frame, items))])
            ARTIFACT_ROOT.mkdir(parents=True, exist_ok=True)
            created = datetime.now()
            path = ARTIFACT_ROOT / f"{code}_{created:%Y%m%d%H%M%S}.xlsx"
            path.write_bytes(data)
            st.session_state["artifact"] = (code, created, data, len(frame), path.name)
        except Exception as error:
            st.error(f"산출물 생성 실패: {error}", icon=":material/error:")
    payload = st.session_state.get("artifact")
    if payload and payload[0] == code:
        st.success(f"{payload[3]:,}건 생성: {payload[4]}", icon=":material/check_circle:")
        st.download_button("엑셀 다운로드", payload[2], file_name=f"{LAYOUTS[code]['name']}_{payload[1]:%Y%m%d%H%M%S}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", icon=":material/download:")
