from __future__ import annotations

import math
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from datetime import datetime
from io import BytesIO
from time import perf_counter
from typing import Any

import oracledb
import pandas as pd
import streamlit as st


DEFAULT_LAYOUT_TABLE = "TB_TABLE_LAYOUT_TLP"
MIGRATION_MAPPING_TABLE = "TB_MIG_TABLE_INFO"
DEFAULT_TARGET_OWNER = "PCERP_RENTALAPP_MIG"
DEFAULT_TABLESPACE = "PCERP_MIG_DATA"
DEFAULT_INDEX_TABLESPACE = "PCERP_MIG_INEX"
DDL_PREVIEW_MAX_LINES = 500
ORACLE_RESERVED_WORDS = {
    "ACCESS", "ADD", "ALL", "ALTER", "AND", "ANY", "AS", "ASC", "AUDIT", "BETWEEN",
    "BY", "CHAR", "CHECK", "CLUSTER", "COLUMN", "COMMENT", "COMPRESS", "CONNECT", "CREATE",
    "CURRENT", "DATE", "DECIMAL", "DEFAULT", "DELETE", "DESC", "DISTINCT", "DROP", "ELSE",
    "EXCLUSIVE", "EXISTS", "FILE", "FLOAT", "FOR", "FROM", "GRANT", "GROUP", "HAVING", "IDENTIFIED",
    "IMMEDIATE", "IN", "INCREMENT", "INDEX", "INITIAL", "INSERT", "INTEGER", "INTERSECT", "INTO",
    "IS", "LEVEL", "LIKE", "LOCK", "LONG", "MAXEXTENTS", "MINUS", "MLSLABEL", "MODE", "MODIFY",
    "NOAUDIT", "NOCOMPRESS", "NOT", "NOWAIT", "NULL", "NUMBER", "OF", "OFFLINE", "ON", "ONLINE",
    "OPTION", "OR", "ORDER", "PCTFREE", "PRIOR", "PRIVILEGES", "PUBLIC", "RAW", "RENAME", "RESOURCE",
    "REVOKE", "ROW", "ROWID", "ROWLABEL", "ROWNUM", "ROWS", "SELECT", "SESSION", "SET", "SHARE",
    "SIZE", "SMALLINT", "START", "SUCCESSFUL", "SYNONYM", "SYSDATE", "TABLE", "THEN", "TO", "TRIGGER",
    "UID", "UNION", "UNIQUE", "UPDATE", "USER", "VALIDATE", "VALUES", "VARCHAR", "VARCHAR2", "VIEW",
    "WHENEVER", "WHERE", "WITH",
}


@dataclass(frozen=True)
class DdlStatement:
    action: str
    source_db: str
    source_table: str
    target_table: str
    sql: str


@dataclass(frozen=True)
class GeneratedDdl:
    statements: tuple[DdlStatement, ...]
    text: str
    structure_text: str
    comment_text: str
    generated_at: str
    table_count: int
    comment_count: int
    source_keys: tuple[tuple[str, str], ...]


def init_state() -> None:
    defaults: dict[str, Any] = {
        "comparison": None,
        "ddl_artifact": None,
        "ddl_logs": None,
        "selected_owner": None,
        "migration_sync_logs": None,
    }
    for key, value in defaults.items():
        st.session_state.setdefault(key, value)


def sql_name(value: str) -> str:
    name = str(value or "").strip().upper()
    if not re.fullmatch(r"[A-Z][A-Z0-9_$#]{0,127}", name):
        raise ValueError(f"Oracle 객체명으로 사용할 수 없습니다: {value}")
    return name


def clean_text(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def normalized(value: object) -> str:
    return re.sub(r"\s+", " ", clean_text(value)).upper()


def normalized_series(values: pd.Series) -> pd.Series:
    return values.astype("string").fillna("").str.strip().str.replace(r"\s+", " ", regex=True).str.upper()


def comparison_key(value: object) -> str:
    text = clean_text(value)
    if len(text) >= 2 and ((text.startswith('"') and text.endswith('"')) or (text.startswith("[") and text.endswith("]"))):
        text = text[1:-1]
    return re.sub(r"\s+", " ", text).upper()


def snapshot_label(value: object) -> str:
    raw = clean_text(value)
    digits = re.sub(r"\D", "", raw)
    if len(digits) == 8:
        return f"{digits[:4]}-{digits[4:6]}-{digits[6:]}"
    return raw


def colno_number(value: object) -> int:
    try:
        return int(float(clean_text(value)))
    except (TypeError, ValueError):
        return 999999999


def parse_length(value: object) -> tuple[int | None, int | None, bool]:
    raw = clean_text(value).replace(" ", "")
    if not raw or raw.upper() == "MAX":
        return None, None, raw.upper() == "MAX"
    match = re.fullmatch(r"(\d+)(?:,(\d+))?", raw)
    if not match:
        return None, None, False
    return int(match.group(1)), int(match.group(2)) if match.group(2) else None, False


def q_literal(value: object) -> str:
    text = re.sub(r"[\r\n]+", " ", clean_text(value)).strip()[:3900]
    for opening, closing in (("[", "]"), ("{", "}"), ("(", ")"), ("<", ">")):
        if closing not in text:
            return f"q'{opening}{text}{closing}'"
    return "'" + text.replace("'", "''") + "'"


def source_column_name(value: object, colno: object, used: set[str]) -> str:
    raw = re.sub(r"\s+", "", clean_text(value).strip("'").upper())
    cleaned = re.sub(r"[^A-Za-z0-9_$\#ㄱ-ㅎㅏ-ㅣ가-힣]", "_", raw)
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    requires_prefix = not raw or not raw[0].isalpha()
    base = raw if not requires_prefix else f"C{cleaned}"
    base = base[:30]
    candidate = base
    suffix = f"_{colno_number(colno)}"
    if candidate in used:
        candidate = f"{base[:30 - len(suffix)]}{suffix}"
    sequence = 2
    while candidate in used:
        numbered = f"_{sequence}"
        candidate = f"{base[:30 - len(suffix) - len(numbered)]}{suffix}{numbered}"
        sequence += 1
    used.add(candidate)
    return candidate


def oracle_column_identifier(value: str) -> str:
    return f'"{value}"' if value in ORACLE_RESERVED_WORDS else value


def target_column_note(source_column: object, target_column: str) -> str:
    source_name = clean_text(source_column).strip("'").upper()
    notes: list[str] = []
    if target_column != source_name:
        notes.append("☆컬럼변경")
    if source_name in ORACLE_RESERVED_WORDS:
        notes.append("◎예약어")
    return " · ".join(notes)


def fallback_target_table(owner: str, table: str) -> str:
    source = f"{normalized(owner)}_{normalized(table)}"
    source = re.sub(r"[^A-Za-z0-9_$\#ㄱ-ㅎㅏ-ㅣ가-힣]", "_", source)
    source = re.sub(r"_+", "_", source).strip("_")
    if not source or not source[0].isalpha():
        source = f"T_{source}"
    return source[:128]


def target_table_name(owner: str, table: str) -> str:
    return fallback_target_table(owner, table)


def migration_target_tables(table_changes: pd.DataFrame) -> list[str]:
    return sorted(
        {
            target_table_name(normalized(row["DB"]), normalized(row["테이블"]))
            for _, row in table_changes.iterrows()
        }
    )


def is_not_null(value: object) -> bool:
    return normalized(value) in {"N", "NO", "NOT NULL", "NOT_NULL", "0", "FALSE"}


def is_primary_key(value: object) -> bool:
    return normalized(value) in {"Y", "YES", "PK", "P", "1", "TRUE"}


def mssql_to_oracle(data_type: object, length: object, character_multiplier: int) -> str:
    data_type_text = normalized(data_type).lower()
    precision, scale, is_max = parse_length(length)
    char_types = {"varchar", "varchar2", "nvarchar", "nvarchar2", "char", "nchar"}
    if data_type_text in char_types:
        if is_max or precision is None or precision <= 0:
            return "CLOB"
        converted_length = math.ceil(precision * character_multiplier)
        return "CLOB" if converted_length > 4000 else f"VARCHAR2({converted_length})"
    if data_type_text in {"text", "ntext", "xml", "sql_variant"}:
        return "CLOB"
    if data_type_text in {"tinyint", "smallint", "int", "integer", "bigint"}:
        digits = {"tinyint": 3, "smallint": 5, "int": 10, "integer": 10, "bigint": 19}[data_type_text]
        return f"NUMBER({digits})"
    if data_type_text in {"bit", "boolean"}:
        return "NUMBER(1)"
    if data_type_text in {"decimal", "numeric", "number"}:
        if precision is not None and scale is not None:
            return f"NUMBER({precision},{scale})"
        if precision is not None:
            return f"NUMBER({precision})"
        return "NUMBER"
    if data_type_text in {"money", "smallmoney"}:
        return "NUMBER(19,4)"
    if data_type_text in {"float", "real"}:
        return "NUMBER(38,17)"
    if data_type_text in {"datetime", "datetime2", "smalldatetime", "date"}:
        return "TIMESTAMP"
    if data_type_text in {"datetimeoffset"}:
        return "TIMESTAMP WITH TIME ZONE"
    if data_type_text in {"time"}:
        return "VARCHAR2(30)"
    if data_type_text in {"binary", "varbinary", "image", "rowversion", "timestamp"}:
        return "BLOB"
    if data_type_text in {"uniqueidentifier", "guid"}:
        return "VARCHAR2(36)"
    return "VARCHAR2(4000)"


@st.cache_resource(show_spinner=False)
def get_pool() -> oracledb.ConnectionPool:
    conf = st.secrets["oracle"]
    dsn = f"{conf['host']}:{conf['port']}/{conf['service_name']}"
    return oracledb.create_pool(
        user=conf["user"],
        password=conf["password"],
        dsn=dsn,
        min=1,
        max=4,
        increment=1,
    )


def dataframe_from_cursor(cursor: oracledb.Cursor) -> pd.DataFrame:
    columns = [description[0].upper() for description in cursor.description]
    return pd.DataFrame(cursor.fetchall(), columns=columns)


@st.cache_data(ttl=120, show_spinner=False)
def fetch_snapshot_dates(layout_owner: str, layout_table: str) -> list[str]:
    owner = sql_name(layout_owner)
    table = sql_name(layout_table)
    with get_pool().acquire() as conn:
        with conn.cursor() as cursor:
            cursor.execute(f"SELECT DISTINCT STD_DT FROM {owner}.{table} ORDER BY STD_DT")
            return [clean_text(row[0]) for row in cursor.fetchall() if clean_text(row[0])]


@st.cache_data(ttl=120, max_entries=20, show_spinner=False)
def fetch_layout_pair(
    layout_owner: str,
    layout_table: str,
    before_snapshot: str,
    after_snapshot: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    owner = sql_name(layout_owner)
    table = sql_name(layout_table)
    query = f"""
        SELECT STD_DT, OWNER, TBL, ENTITY, COLNO, COL, ATTR, DATATYPE, LEN, ISPK, NULLABLE
          FROM {owner}.{table}
         WHERE STD_DT IN (:before_snapshot, :after_snapshot)
         ORDER BY OWNER, TBL, COLNO
    """
    with get_pool().acquire() as conn:
        with conn.cursor() as cursor:
            cursor.execute(query, {"before_snapshot": before_snapshot, "after_snapshot": after_snapshot})
            frame = dataframe_from_cursor(cursor)
    return (
        frame[frame["STD_DT"].map(clean_text) == before_snapshot].copy(),
        frame[frame["STD_DT"].map(clean_text) == after_snapshot].copy(),
    )




def table_signature(table_frame: pd.DataFrame) -> tuple[tuple[str, ...], ...]:
    signature_columns = ["_SIG_COLNO", "_SIG_COL", "_SIG_ATTR", "_SIG_DATATYPE", "_SIG_LEN", "_SIG_ISPK", "_SIG_NULLABLE"]
    ordered = table_frame.sort_values("_COLNO_ORDER")
    return tuple(ordered[signature_columns].itertuples(index=False, name=None))


def table_groups(frame: pd.DataFrame) -> dict[tuple[str, str], pd.DataFrame]:
    prepared = frame.copy()
    prepared["_COMPARE_OWNER"] = prepared["OWNER"].map(comparison_key)
    prepared["_COMPARE_TABLE"] = prepared["TBL"].map(comparison_key)
    prepared["_COLNO_ORDER"] = pd.to_numeric(prepared["COLNO"], errors="coerce").fillna(999999999)
    for column in ("COLNO", "COL", "ATTR", "DATATYPE", "LEN", "ISPK", "NULLABLE"):
        prepared[f"_SIG_{column}"] = normalized_series(prepared[column])
    groups: dict[tuple[str, str], pd.DataFrame] = {}
    for (owner, table), group in prepared.groupby(["_COMPARE_OWNER", "_COMPARE_TABLE"], dropna=False, sort=False):
        groups[(owner, table)] = group.copy()
    return groups


def first_entity(frame: pd.DataFrame) -> str:
    if frame.empty:
        return ""
    values = [clean_text(value) for value in frame["ENTITY"].tolist() if clean_text(value)]
    return values[0] if values else ""


def compare_layouts(before: pd.DataFrame, after: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    before_groups = table_groups(before)
    after_groups = table_groups(after)
    table_rows: list[dict[str, str]] = []
    column_rows: list[dict[str, str]] = []
    for owner, table in sorted(set(before_groups) | set(after_groups)):
        before_frame = before_groups.get((owner, table), pd.DataFrame(columns=before.columns))
        after_frame = after_groups.get((owner, table), pd.DataFrame(columns=after.columns))
        if before_frame.empty:
            status = "신규"
        elif after_frame.empty:
            status = "삭제"
        elif table_signature(before_frame) != table_signature(after_frame) or first_entity(before_frame) != first_entity(after_frame):
            status = "변경"
        else:
            continue
        before_entity = first_entity(before_frame)
        after_entity = first_entity(after_frame)
        entity = after_entity or before_entity
        table_rows.append(
            {
                "DB": owner,
                "테이블": table,
                "엔티티(전)": before_entity,
                "엔티티(후)": after_entity,
                "구분": status,
            }
        )
        before_by_colno = {normalized(row["COLNO"]): row for _, row in before_frame.iterrows()}
        after_by_colno = {normalized(row["COLNO"]): row for _, row in after_frame.iterrows()}
        for colno in sorted(set(before_by_colno) | set(after_by_colno), key=colno_number):
            previous = before_by_colno.get(colno)
            current = after_by_colno.get(colno)
            field_names = ("COL", "ATTR", "DATATYPE", "LEN", "ISPK", "NULLABLE")
            if previous is not None and current is not None and all(normalized(previous[name]) == normalized(current[name]) for name in field_names):
                continue
            column_status = "신규" if previous is None else "삭제" if current is None else "변경"
            column_rows.append(
                {
                    "DB": owner,
                    "테이블": table,
                    "엔티티": entity,
                    "구분": column_status,
                    "컬럼순서": colno,
                    "컬럼명(전)": clean_text(previous["COL"]) if previous is not None else "",
                    "컬럼명(후)": clean_text(current["COL"]) if current is not None else "",
                    "속성(전)": clean_text(previous["ATTR"]) if previous is not None else "",
                    "속성(후)": clean_text(current["ATTR"]) if current is not None else "",
                    "타입(전)": clean_text(previous["DATATYPE"]) if previous is not None else "",
                    "타입(후)": clean_text(current["DATATYPE"]) if current is not None else "",
                    "길이(전)": clean_text(previous["LEN"]) if previous is not None else "",
                    "길이(후)": clean_text(current["LEN"]) if current is not None else "",
                    "PK(전)": clean_text(previous["ISPK"]) if previous is not None else "",
                    "PK(후)": clean_text(current["ISPK"]) if current is not None else "",
                    "NOT NULL(전)": "Y" if previous is not None and is_not_null(previous["NULLABLE"]) else "",
                    "NOT NULL(후)": "Y" if current is not None and is_not_null(current["NULLABLE"]) else "",
                }
            )
    return pd.DataFrame(table_rows), pd.DataFrame(column_rows)


def storage_clause(allocated_bytes: int | None) -> str:
    if allocated_bytes is None or allocated_bytes <= 0:
        return ""
    for divisor, suffix in ((1024**3, "G"), (1024**2, "M"), (1024, "K")):
        if allocated_bytes >= divisor:
            return f" STORAGE (INITIAL {math.ceil(allocated_bytes / divisor)}{suffix})"
    return " STORAGE (INITIAL 1K)"


def clean_metadata_ddl(text: str) -> str:
    value = text.read() if hasattr(text, "read") else str(text)
    value = re.sub(r"\s+", " ", value).strip()
    return value[:-1].strip() if value.endswith(";") else value


def fetch_current_table_allocations(target_owner: str, target_tables: list[str]) -> dict[str, int]:
    if not target_tables:
        return {}
    result: dict[str, int] = {}
    for start in range(0, len(target_tables), 900):
        current_tables = target_tables[start : start + 900]
        bind_names = [f"table_{index}" for index in range(len(current_tables))]
        bind_sql = ", ".join(f":{name}" for name in bind_names)
        binds: dict[str, str] = dict(zip(bind_names, current_tables))
        query = f"""
        SELECT SEGMENT_NAME AS TABLE_NAME, SUM(BYTES) AS ALLOCATED_BYTES
          FROM USER_SEGMENTS
         WHERE SEGMENT_TYPE = 'TABLE'
           AND SEGMENT_NAME IN ({bind_sql})
         GROUP BY SEGMENT_NAME
    """
        try:
            with get_pool().acquire() as conn:
                with conn.cursor() as cursor:
                    cursor.execute(query, binds)
                    result.update({normalized(row[0]): int(row[1]) for row in cursor.fetchall() if row[1] is not None})
        except Exception as exc:
            raise RuntimeError(
                f"Oracle 현재 테이블 할당량 조회 실패: USER_SEGMENTS 조회 오류입니다. 원본 오류: {exc}"
            ) from exc
    return result


def size_label(allocated_bytes: int | float | None) -> str:
    value = int(allocated_bytes or 0)
    if value <= 0:
        return ""
    for divisor, suffix in ((1024**3, "GB"), (1024**2, "MB"), (1024, "KB")):
        if value >= divisor:
            return f"{value / divisor:.2f}".rstrip("0").rstrip(".") + suffix
    return f"{value}B"


def thousand_number_columns(frame: pd.DataFrame) -> dict[str, st.column_config.NumberColumn]:
    return {
        column: st.column_config.NumberColumn(format="localized")
        for column in frame.select_dtypes(include="number").columns
    }


@st.cache_data(ttl=120, max_entries=20, show_spinner=False)
def fetch_table_catalog_metrics(target_tables: tuple[str, ...]) -> pd.DataFrame:
    if not target_tables:
        return pd.DataFrame(columns=["대상 Oracle 테이블", "NUM_ROWS", "INSERTS", "DELETES", "할당 크기"])
    rows: list[dict[str, object]] = []
    for start in range(0, len(target_tables), 900):
        names = target_tables[start : start + 900]
        bind_names = [f"table_{index}" for index in range(len(names))]
        bind_sql = ", ".join(f":{name}" for name in bind_names)
        query = f"""
            SELECT t.table_name,
                   t.num_rows,
                   NVL(m.inserts, 0) AS inserts,
                   NVL(m.deletes, 0) AS deletes,
                   NVL(s.allocated_bytes, 0) AS allocated_bytes
              FROM user_tables t
              LEFT JOIN user_tab_modifications m
                ON m.table_name = t.table_name
              LEFT JOIN (
                    SELECT segment_name, SUM(bytes) AS allocated_bytes
                      FROM user_segments
                     WHERE segment_type = 'TABLE'
                     GROUP BY segment_name
              ) s
                ON s.segment_name = t.table_name
             WHERE t.table_name IN ({bind_sql})
             ORDER BY t.table_name
        """
        with get_pool().acquire() as conn:
            with conn.cursor() as cursor:
                cursor.execute(query, dict(zip(bind_names, names)))
                for table_name, num_rows, inserts, deletes, allocated_bytes in cursor.fetchall():
                    rows.append(
                        {
                            "대상 Oracle 테이블": clean_text(table_name),
                            "NUM_ROWS": int(num_rows) if num_rows is not None else None,
                            "INSERTS": int(inserts or 0),
                            "DELETES": int(deletes or 0),
                            "할당 크기": size_label(allocated_bytes),
                        }
                    )
    return pd.DataFrame(rows, columns=["대상 Oracle 테이블", "NUM_ROWS", "INSERTS", "DELETES", "할당 크기"])


def add_table_catalog_metrics(table_changes: pd.DataFrame) -> pd.DataFrame:
    if table_changes.empty:
        return table_changes
    enriched = table_changes.copy()
    enriched["대상 Oracle 테이블"] = enriched.apply(
        lambda row: target_table_name(normalized(row["DB"]), normalized(row["테이블"])), axis=1
    )
    metrics = fetch_table_catalog_metrics(tuple(sorted(enriched["대상 Oracle 테이블"].unique())))
    return enriched.merge(metrics, on="대상 Oracle 테이블", how="left")


def fetch_index_allocations(cursor: oracledb.Cursor, index_names: list[str]) -> dict[str, int]:
    if not index_names:
        return {}
    allocations: dict[str, int] = {}
    for start in range(0, len(index_names), 900):
        names = index_names[start : start + 900]
        bind_names = [f"index_{index}" for index in range(len(names))]
        bind_sql = ", ".join(f":{name}" for name in bind_names)
        cursor.execute(
            f"""
            SELECT SEGMENT_NAME, SUM(BYTES) AS ALLOCATED_BYTES
              FROM USER_SEGMENTS
             WHERE SEGMENT_TYPE = 'INDEX'
               AND SEGMENT_NAME IN ({bind_sql})
             GROUP BY SEGMENT_NAME
            """,
            dict(zip(bind_names, names)),
        )
        allocations.update({normalized(name): int(size) for name, size in cursor.fetchall() if size is not None})
    return allocations


def apply_storage_allocation(ddl: str, allocated_bytes: int | None) -> str:
    clause = storage_clause(allocated_bytes)
    if not clause:
        return ddl
    if re.search(r"\bSTORAGE\s*\([^()]*\)", ddl, flags=re.IGNORECASE):
        return re.sub(r"\s*STORAGE\s*\([^()]*\)", clause, ddl, count=1, flags=re.IGNORECASE)
    tablespace_match = re.search(r"\s+TABLESPACE\b", ddl, flags=re.IGNORECASE)
    if tablespace_match:
        return f"{ddl[:tablespace_match.start()]}{clause}{ddl[tablespace_match.start():]}"
    return f"{ddl}{clause}"


def existing_index_statements(cursor: oracledb.Cursor, target_owner: str, target_table: str) -> list[str]:
    query = """
        SELECT i.index_name
          FROM user_indexes i
         WHERE i.table_name = :table_name
           AND i.index_type NOT IN ('LOB', 'IOT - TOP')
           AND i.index_name NOT LIKE 'SYS$_%' ESCAPE '$'
           AND NOT EXISTS (
                SELECT 1
                  FROM user_constraints c
                 WHERE c.table_name = i.table_name
                   AND c.constraint_type IN ('P', 'U')
                   AND c.index_name = i.index_name
           )
         ORDER BY i.index_name
    """
    cursor.execute(query, {"table_name": target_table})
    index_names = [clean_text(row[0]) for row in cursor.fetchall()]
    index_allocations = fetch_index_allocations(cursor, index_names)
    statements: list[str] = []
    for index_name in index_names:
        cursor.execute(
            "SELECT DBMS_METADATA.GET_DDL('INDEX', :index_name, :index_owner) FROM DUAL",
            {"index_name": index_name, "index_owner": target_owner},
        )
        row = cursor.fetchone()
        if row and row[0]:
            statements.append(apply_storage_allocation(clean_metadata_ddl(row[0]), index_allocations.get(normalized(index_name))))
    return statements


def existing_constraint_statements(cursor: oracledb.Cursor, target_owner: str, target_table: str) -> list[str]:
    cursor.execute(
        """
        SELECT constraint_name, index_name
          FROM user_constraints
         WHERE table_name = :table_name
           AND constraint_type IN ('P', 'U')
         ORDER BY constraint_name
        """,
        {"table_name": target_table},
    )
    constraints = [(clean_text(name), clean_text(index_name)) for name, index_name in cursor.fetchall()]
    index_allocations = fetch_index_allocations(cursor, [index_name for _, index_name in constraints if index_name])
    statements: list[str] = []
    for constraint_name, index_name in constraints:
        cursor.execute(
            "SELECT DBMS_METADATA.GET_DDL('CONSTRAINT', :constraint_name, :owner) FROM DUAL",
            {"constraint_name": constraint_name, "owner": target_owner},
        )
        row = cursor.fetchone()
        if row and row[0]:
            statements.append(apply_storage_allocation(clean_metadata_ddl(row[0]), index_allocations.get(normalized(index_name))))
    return statements


def build_table_ddl(
    target_owner: str,
    target_table: str,
    owner: str,
    source_table: str,
    entity: str,
    layout: pd.DataFrame,
    table_tablespace: str,
    index_tablespace: str,
    character_multiplier: int,
    initial_size_bytes: int | None,
) -> tuple[str, list[str]]:
    used_columns: set[str] = set()
    column_lines: list[str] = []
    column_comments: list[str] = []
    for _, row in layout.sort_values("COLNO", key=lambda item: item.map(colno_number)).iterrows():
        source_name = clean_text(row["COL"])
        source_attr = clean_text(row["ATTR"])
        oracle_name = source_column_name(source_name, row["COLNO"], used_columns)
        oracle_identifier = oracle_column_identifier(oracle_name)
        oracle_type = mssql_to_oracle(row["DATATYPE"], row["LEN"], character_multiplier)
        column_lines.append(f"    {oracle_identifier} {oracle_type}")
        if source_attr:
            column_comments.append(f"COMMENT ON COLUMN {target_owner}.{target_table}.{oracle_identifier} IS {q_literal(source_attr)}")
    ddl = f"CREATE TABLE {target_owner}.{target_table} (\n"
    ddl += ",\n".join(column_lines)
    ddl += f"\n) SEGMENT CREATION IMMEDIATE NOLOGGING TABLESPACE {table_tablespace}{storage_clause(initial_size_bytes)}"
    table_comment = clean_text(entity)
    comments = ([] if not table_comment else [f"COMMENT ON TABLE {target_owner}.{target_table} IS {q_literal(table_comment)}"]) + column_comments
    return ddl, comments


def build_ddl_artifact(
    table_changes: pd.DataFrame,
    before_layout: pd.DataFrame,
    after_layout: pd.DataFrame,
    target_owner: str,
    table_tablespace: str,
    index_tablespace: str,
    character_multiplier: int,
) -> GeneratedDdl:
    before_groups = table_groups(before_layout)
    after_groups = table_groups(after_layout)
    statements: list[DdlStatement] = []
    existing_tables = [
        target_table_name(normalized(change["DB"]), normalized(change["테이블"]))
        for _, change in table_changes.iterrows()
        if clean_text(change["구분"]) == "변경"
    ]
    current_allocations = fetch_current_table_allocations(target_owner, sorted(set(existing_tables)))
    with get_pool().acquire() as conn:
        with conn.cursor() as cursor:
            for _, change in table_changes.iterrows():
                source_owner = normalized(change["DB"])
                source_table = normalized(change["테이블"])
                status = clean_text(change["구분"])
                source_layout = after_groups.get((source_owner, source_table))
                target_table = target_table_name(source_owner, source_table)
                if status in {"변경", "삭제"}:
                    statements.append(
                        DdlStatement(
                            "DROP",
                            source_owner,
                            source_table,
                            target_table,
                            f"DROP TABLE {target_owner}.{target_table} CASCADE CONSTRAINTS PURGE",
                        )
                    )
                if status == "삭제" or source_layout is None or source_layout.empty:
                    continue
                entity = first_entity(source_layout)
                create_sql, comment_sqls = build_table_ddl(
                    target_owner,
                    target_table,
                    source_owner,
                    source_table,
                    entity,
                    source_layout,
                    table_tablespace,
                    index_tablespace,
                    character_multiplier,
                    current_allocations.get(target_table) if status == "변경" else None,
                )
                statements.append(DdlStatement("CREATE TABLE", source_owner, source_table, target_table, create_sql))
                for comment_sql in comment_sqls:
                    statements.append(DdlStatement("COMMENT", source_owner, source_table, target_table, comment_sql))
                if status == "변경":
                    try:
                        constraint_statements = existing_constraint_statements(cursor, target_owner, target_table)
                    except Exception as exc:
                        raise RuntimeError(
                            f"기존 PK·UNIQUE 제약 추출 실패: {target_owner}.{target_table}. USER_CONSTRAINTS 또는 DBMS_METADATA 조회 오류입니다. 원본 오류: {exc}"
                        ) from exc
                    try:
                        index_statements = existing_index_statements(cursor, target_owner, target_table)
                    except Exception as exc:
                        raise RuntimeError(
                            f"기존 인덱스 추출 실패: {target_owner}.{target_table}. USER_INDEXES 또는 DBMS_METADATA 조회 오류입니다. 원본 오류: {exc}"
                        ) from exc
                    for constraint_sql in constraint_statements:
                        statements.append(DdlStatement("CREATE CONSTRAINT", source_owner, source_table, target_table, constraint_sql))
                    for index_sql in index_statements:
                        statements.append(DdlStatement("CREATE INDEX", source_owner, source_table, target_table, index_sql))
    def statement_text(items: list[DdlStatement]) -> str:
        blocks: list[str] = []
        current_key: tuple[str, str] | None = None
        current_lines: list[str] = []
        for statement in items:
            statement_key = (statement.source_db, statement.source_table)
            if current_key is not None and statement_key != current_key:
                blocks.append("\n".join(current_lines))
                current_lines = []
            current_key = statement_key
            current_lines.append(f"{statement.sql};")
        if current_lines:
            blocks.append("\n".join(current_lines))
        return "\n\n".join(blocks) + ("\n" if blocks else "")

    text = statement_text(statements)
    structure_text = statement_text([statement for statement in statements if statement.action != "COMMENT"])
    comment_statements = [statement for statement in statements if statement.action == "COMMENT"]
    comment_text = statement_text(comment_statements)
    source_keys = tuple(sorted({(statement.source_db, statement.source_table) for statement in statements}))
    return GeneratedDdl(
        tuple(statements),
        text,
        structure_text,
        comment_text,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        len(source_keys),
        len(comment_statements),
        source_keys,
    )


def generated_table_grid(artifact: GeneratedDdl, comment_only: bool = False) -> pd.DataFrame:
    rows: list[dict[str, str]] = []
    for source_db, source_table in artifact.source_keys:
        statements = [item for item in artifact.statements if (item.source_db, item.source_table) == (source_db, source_table)]
        if comment_only:
            statements = [item for item in statements if item.action == "COMMENT"]
        if not statements:
            continue
        target_table = statements[0].target_table if statements else ""
        rows.append({"DB": source_db, "테이블": source_table, "대상 Oracle 테이블": target_table, "DDL 구문 수": len(statements)})
    return pd.DataFrame(rows)


def generated_column_mapping_grid(
    table_changes: pd.DataFrame,
    after_layout: pd.DataFrame,
    character_multiplier: int,
) -> pd.DataFrame:
    columns = [
        "SRC SYSTEM",
        "SRC TBL",
        "SRC ENT",
        "SRC COLNO",
        "SRC COL",
        "SRC ATTR",
        "SRC DATATYPE",
        "SRC LEN",
        "SRC ISPK",
        "SRC NULLABLE",
        "TGT COL",
        "TGT DATATYPE",
        "TGT 비고",
    ]
    after_groups = table_groups(after_layout)
    rows: list[dict[str, object]] = []
    for _, table_change in table_changes.iterrows():
        source_system = normalized(table_change["DB"])
        source_table = normalized(table_change["테이블"])
        layout = after_groups.get((source_system, source_table))
        if layout is None or layout.empty:
            continue
        used_columns: set[str] = set()
        source_entity = first_entity(layout)
        for _, row in layout.sort_values("COLNO", key=lambda item: item.map(colno_number)).iterrows():
            target_column = source_column_name(row["COL"], row["COLNO"], used_columns)
            rows.append(
                {
                    "SRC SYSTEM": source_system,
                    "SRC TBL": source_table,
                    "SRC ENT": source_entity,
                    "SRC COLNO": clean_text(row["COLNO"]),
                    "SRC COL": clean_text(row["COL"]),
                    "SRC ATTR": clean_text(row["ATTR"]),
                    "SRC DATATYPE": clean_text(row["DATATYPE"]),
                    "SRC LEN": clean_text(row["LEN"]),
                    "SRC ISPK": clean_text(row["ISPK"]),
                    "SRC NULLABLE": clean_text(row["NULLABLE"]),
                    "TGT COL": target_column,
                    "TGT DATATYPE": mssql_to_oracle(row["DATATYPE"], row["LEN"], character_multiplier),
                    "TGT 비고": target_column_note(row["COL"], target_column),
                }
            )
    return pd.DataFrame(rows, columns=columns)


def column_mapping_excel_bytes(frame: pd.DataFrame) -> bytes:
    from openpyxl.styles import Border, Font, PatternFill, Side

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name="컬럼매핑", index=False)
        worksheet = writer.sheets["컬럼매핑"]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        thin_side = Side(style="thin", color="9EADBF")
        all_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        header_font = Font(name="맑은 고딕", size=10, bold=True)
        body_font = Font(name="맑은 고딕", size=10)
        header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        for row_index, row in enumerate(worksheet.iter_rows(), start=1):
            for cell in row:
                cell.font = header_font if row_index == 1 else body_font
                cell.border = all_border
                if row_index == 1:
                    cell.fill = header_fill
        for column_cells in worksheet.columns:
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 42)
    return output.getvalue()


def fetch_migration_mapping_rows(target_tables: list[str]) -> pd.DataFrame:
    columns = ["SRC_SYSTEM", "SRC_TABLE", "SRC_ENTITY", "SRC_COLNO", "SRC_COLUMN", "TGT_TABLE", "TGT_COLUMN", "JB_GRP", "MIG_YN", "KTR"]
    if not target_tables:
        return pd.DataFrame(columns=columns)
    rows: list[tuple[object, ...]] = []
    owner = sql_name(DEFAULT_TARGET_OWNER)
    table = sql_name(MIGRATION_MAPPING_TABLE)
    with get_pool().acquire() as conn:
        with conn.cursor() as cursor:
            for start in range(0, len(target_tables), 900):
                names = target_tables[start : start + 900]
                bind_names = [f"table_{index}" for index in range(len(names))]
                bind_sql = ", ".join(f":{name}" for name in bind_names)
                cursor.execute(
                    f"""
                    SELECT SRC_SYSTEM, SRC_TABLE, SRC_ENTITY, SRC_COLNO, SRC_COLUMN,
                           TGT_TABLE, TGT_COLUMN, JB_GRP, MIG_YN, KTR
                      FROM {owner}.{table}
                     WHERE TGT_TABLE IN ({bind_sql})
                    """,
                    dict(zip(bind_names, names)),
                )
                rows.extend(cursor.fetchall())
    result = pd.DataFrame(rows, columns=columns)
    return result.sort_values(["TGT_TABLE", "TGT_COLUMN", "SRC_COLNO"]).reset_index(drop=True) if not result.empty else result


def jb_group_number(value: object) -> Decimal:
    try:
        return Decimal(clean_text(value))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"JB_GRP 숫자값이 올바르지 않습니다: {value}") from exc


def existing_mapping_defaults(existing_rows: pd.DataFrame) -> tuple[dict[str, Decimal], dict[tuple[str, str], str], dict[tuple[str, str], str]]:
    jb_groups: dict[str, Decimal] = {}
    mig_yn_by_column: dict[tuple[str, str], str] = {}
    ktr_by_column: dict[tuple[str, str], str] = {}
    if existing_rows.empty:
        return jb_groups, mig_yn_by_column, ktr_by_column
    for target_table, group in existing_rows.groupby("TGT_TABLE", sort=False):
        group_values = {jb_group_number(value) for value in group["JB_GRP"].tolist() if clean_text(value)}
        if len(group_values) > 1:
            raise ValueError(f"TB_MIG_TABLE_INFO의 {target_table} JB_GRP 값이 여러 개입니다: {', '.join(str(value) for value in sorted(group_values))}")
        if group_values:
            jb_groups[normalized(target_table)] = next(iter(group_values))
    for _, row in existing_rows.iterrows():
        key = (normalized(row["TGT_TABLE"]), normalized(row["TGT_COLUMN"]))
        value = normalized(row["MIG_YN"])
        previous = mig_yn_by_column.get(key)
        if previous is not None and previous != value:
            raise ValueError(f"TB_MIG_TABLE_INFO의 {key[0]}.{key[1]} MIG_YN 값이 일치하지 않습니다.")
        mig_yn_by_column[key] = value
        ktr_value = clean_text(row["KTR"])
        previous_ktr = ktr_by_column.get(key)
        if previous_ktr is not None and previous_ktr != ktr_value:
            raise ValueError(f"TB_MIG_TABLE_INFO의 {key[0]}.{key[1]} KTR 값이 일치하지 않습니다.")
        ktr_by_column[key] = ktr_value
    return jb_groups, mig_yn_by_column, ktr_by_column


def migration_mapping_rows(
    column_mapping: pd.DataFrame,
    jb_groups: dict[str, Decimal],
    mig_yn_by_column: dict[tuple[str, str], str],
    ktr_by_column: dict[tuple[str, str], str],
) -> pd.DataFrame:
    columns = ["SRC_SYSTEM", "SRC_TABLE", "SRC_ENTITY", "SRC_COLNO", "SRC_COLUMN", "TGT_TABLE", "TGT_COLUMN", "JB_GRP", "MIG_YN", "KTR"]
    rows: list[dict[str, object]] = []
    for _, row in column_mapping.iterrows():
        target_table = target_table_name(normalized(row["SRC SYSTEM"]), normalized(row["SRC TBL"]))
        target_column = normalized(row["TGT COL"])
        jb_group = jb_groups.get(target_table)
        if jb_group is None:
            raise ValueError(f"{target_table}의 JB_GRP 값을 입력해 주십시오.")
        colno = colno_number(row["SRC COLNO"])
        if colno == 999999999:
            raise ValueError(f"{target_table}.{target_column}의 SRC_COLNO 값이 올바르지 않습니다.")
        rows.append(
            {
                "SRC_SYSTEM": normalized(row["SRC SYSTEM"]),
                "SRC_TABLE": normalized(row["SRC TBL"]),
                "SRC_ENTITY": clean_text(row["SRC ENT"]),
                "SRC_COLNO": colno,
                "SRC_COLUMN": clean_text(row["SRC COL"]),
                "TGT_TABLE": target_table,
                "TGT_COLUMN": target_column,
                "JB_GRP": jb_group,
                "MIG_YN": mig_yn_by_column.get((target_table, target_column), "N"),
                "KTR": ktr_by_column.get((target_table, target_column), None),
            }
        )
    return pd.DataFrame(rows, columns=columns)


def execute_migration_mapping_sync(target_tables: list[str], mapping_rows: pd.DataFrame) -> pd.DataFrame:
    owner = sql_name(DEFAULT_TARGET_OWNER)
    table = sql_name(MIGRATION_MAPPING_TABLE)
    delete_sql = f"DELETE FROM {owner}.{table} WHERE TGT_TABLE = :target_table"
    insert_sql = f"""
        INSERT INTO {owner}.{table} (
            SRC_SYSTEM, SRC_TABLE, SRC_ENTITY, SRC_COLNO, SRC_COLUMN,
            TGT_TABLE, TGT_COLUMN, JB_GRP, MIG_YN
        ) VALUES (
            :SRC_SYSTEM, :SRC_TABLE, :SRC_ENTITY, :SRC_COLNO, :SRC_COLUMN,
            :TGT_TABLE, :TGT_COLUMN, :JB_GRP, :MIG_YN
        )
    """
    insert_with_ktr_sql = f"""
        INSERT INTO {owner}.{table} (
            SRC_SYSTEM, SRC_TABLE, SRC_ENTITY, SRC_COLNO, SRC_COLUMN,
            TGT_TABLE, TGT_COLUMN, JB_GRP, MIG_YN, KTR
        ) VALUES (
            :SRC_SYSTEM, :SRC_TABLE, :SRC_ENTITY, :SRC_COLNO, :SRC_COLUMN,
            :TGT_TABLE, :TGT_COLUMN, :JB_GRP, :MIG_YN, :KTR
        )
    """
    with get_pool().acquire() as conn:
        try:
            with conn.cursor() as cursor:
                deleted_count = 0
                for target_table in target_tables:
                    cursor.execute(delete_sql, target_table=target_table)
                    deleted_count += cursor.rowcount
                without_ktr = mapping_rows[mapping_rows["KTR"].map(clean_text) == ""]
                with_ktr = mapping_rows[mapping_rows["KTR"].map(clean_text) != ""]
                if not without_ktr.empty:
                    cursor.executemany(insert_sql, without_ktr.drop(columns="KTR").to_dict("records"))
                if not with_ktr.empty:
                    cursor.executemany(insert_with_ktr_sql, with_ktr.to_dict("records"))
                conn.commit()
        except Exception as exc:
            conn.rollback()
            raise RuntimeError(f"TB_MIG_TABLE_INFO 반영에 실패하여 전체 작업을 취소했습니다: {exc}") from exc
    return pd.DataFrame([{"삭제 행 수": deleted_count, "입력 행 수": len(mapping_rows), "대상 테이블 수": len(target_tables)}])


def validate_migration_mapping_rows(mapping_rows: pd.DataFrame, target_tables: list[str]) -> pd.DataFrame:
    columns = ["SRC_SYSTEM", "SRC_TABLE", "SRC_ENTITY", "SRC_COLNO", "SRC_COLUMN", "TGT_TABLE", "TGT_COLUMN", "JB_GRP", "MIG_YN", "KTR"]
    if list(mapping_rows.columns) != columns:
        raise ValueError("TB_MIG_TABLE_INFO 반영 그리드의 컬럼 구성이 올바르지 않습니다.")
    result = mapping_rows.copy()
    for column in ("SRC_SYSTEM", "SRC_TABLE", "TGT_TABLE", "TGT_COLUMN", "MIG_YN"):
        result[column] = result[column].map(normalized)
    for column in ("SRC_ENTITY", "SRC_COLUMN", "KTR"):
        result[column] = result[column].map(clean_text)
    result["SRC_COLNO"] = result["SRC_COLNO"].map(colno_number)
    result["JB_GRP"] = result["JB_GRP"].map(jb_group_number)
    if result["SRC_COLNO"].eq(999999999).any():
        raise ValueError("SRC_COLNO는 숫자로 입력해 주십시오.")
    if not result["TGT_TABLE"].isin(target_tables).all():
        raise ValueError("선택한 TGT_TABLE 범위를 벗어난 행은 반영할 수 없습니다.")
    if result[["SRC_SYSTEM", "SRC_TABLE", "TGT_COLUMN"]].eq("").any().any():
        raise ValueError("SRC_SYSTEM, SRC_TABLE, TGT_COLUMN은 비워 둘 수 없습니다.")
    if not result["MIG_YN"].isin({"Y", "N"}).all():
        raise ValueError("MIG_YN은 Y 또는 N으로 입력해 주십시오.")
    for target_table, group in result.groupby("TGT_TABLE", sort=False):
        if group["JB_GRP"].nunique() != 1:
            raise ValueError(f"{target_table}의 JB_GRP는 하나의 숫자값으로 동일해야 합니다.")
    return result


def render_migration_mapping_sync(table_changes: pd.DataFrame, column_mapping: pd.DataFrame) -> None:
    target_tables = migration_target_tables(table_changes)
    if not target_tables:
        return
    try:
        existing_rows = fetch_migration_mapping_rows(target_tables)
        existing_jb_groups, existing_mig_yn, existing_ktr = existing_mapping_defaults(existing_rows)
    except Exception as exc:
        st.error(str(exc))
        return
    mapping_target_tables = sorted({target_table_name(normalized(row["SRC SYSTEM"]), normalized(row["SRC TBL"])) for _, row in column_mapping.iterrows()})
    jb_groups = existing_jb_groups.copy()
    for target_table in mapping_target_tables:
        jb_groups.setdefault(target_table, Decimal("6"))
    summary = pd.DataFrame(
        [
            {
                "TGT TABLE": target_table,
                "기존 매핑": int((existing_rows["TGT_TABLE"].map(normalized) == target_table).sum()) if not existing_rows.empty else 0,
                "입력 매핑": int(sum(target_table_name(normalized(row["SRC SYSTEM"]), normalized(row["SRC TBL"])) == target_table for _, row in column_mapping.iterrows())),
                "JB_GRP": jb_groups.get(target_table, ""),
            }
            for target_table in target_tables
        ]
    )
    st.caption("선택한 TGT_TABLE의 기존 행을 모두 삭제한 뒤 현재 컬럼 매핑으로 재입력합니다. 기존 JB_GRP·MIG_YN·KTR은 유지하며, 신규 JB_GRP는 숫자 6, 신규 MIG_YN은 N, 신규 KTR은 입력하지 않습니다.")
    st.dataframe(summary, hide_index=True, height=260, column_config=thousand_number_columns(summary))
    try:
        preview_rows = migration_mapping_rows(column_mapping, jb_groups, existing_mig_yn, existing_ktr)
        preview_rows["JB_GRP"] = preview_rows["JB_GRP"].map(float)
    except Exception as exc:
        st.error(str(exc))
        return
    with st.form("migration_mapping_sync_form", border=True):
        edited_rows = st.data_editor(
            preview_rows,
            hide_index=True,
            disabled=["SRC_SYSTEM", "SRC_TABLE", "TGT_TABLE"],
            height=420,
            key="migration_mapping_sync_editor",
            column_config={
                "SRC_COLNO": st.column_config.NumberColumn("SRC_COLNO", min_value=1, step=1, format="%d"),
                "JB_GRP": st.column_config.NumberColumn("JB_GRP", min_value=0, step=0.1, format="%.1f"),
                "MIG_YN": st.column_config.SelectboxColumn("MIG_YN", options=["Y", "N"]),
            },
        )
        confirmed = st.checkbox("표시된 TGT_TABLE의 기존 TB_MIG_TABLE_INFO 행을 삭제하고 재입력하는 것을 확인했습니다.", key="migration_mapping_sync_confirm")
        submitted = st.form_submit_button("TB_MIG_TABLE_INFO 반영", type="primary", icon=":material/save:")
    if not submitted:
        return
    try:
        mapping_rows = validate_migration_mapping_rows(edited_rows, target_tables)
        if not confirmed:
            raise ValueError("삭제 및 재입력 확인을 선택해 주십시오.")
        st.session_state.migration_sync_logs = execute_migration_mapping_sync(target_tables, mapping_rows)
        st.success("TB_MIG_TABLE_INFO 반영을 완료했습니다.", icon=":material/check_circle:")
    except Exception as exc:
        st.error(str(exc))
    if st.session_state.migration_sync_logs is not None:
        st.dataframe(st.session_state.migration_sync_logs, hide_index=True, column_config=thousand_number_columns(st.session_state.migration_sync_logs))


def render_ddl_preview(ddl_text: str) -> None:
    lines = ddl_text.splitlines()
    preview = "\n".join(lines[:DDL_PREVIEW_MAX_LINES])
    if len(lines) > DDL_PREVIEW_MAX_LINES:
        st.warning(
            f"미리보기는 처음 {DDL_PREVIEW_MAX_LINES:,}행만 표시합니다. "
            f"전체 {len(lines):,}행은 다운로드 파일에서 확인해 주세요."
        )
    st.code(preview, language="sql")


def execute_ddl(artifact: GeneratedDdl, selected_keys: set[tuple[str, str]], comment_only: bool = False) -> pd.DataFrame:
    selected_statements = [
        item
        for item in artifact.statements
        if (item.source_db, item.source_table) in selected_keys and (not comment_only or item.action == "COMMENT")
    ]
    status_by_table: dict[tuple[str, str], dict[str, str | int]] = {}
    for statement in selected_statements:
        key = (statement.source_db, statement.source_table)
        status_by_table.setdefault(
            key,
            {"DB": statement.source_db, "테이블": statement.source_table, "대상 Oracle 테이블": statement.target_table, "실행 구문 수": 0, "결과": "성공", "메시지": ""},
        )
    failed_tables: set[tuple[str, str]] = set()
    progress = st.progress(0.0)
    with get_pool().acquire() as conn:
        with conn.cursor() as cursor:
            for number, statement in enumerate(selected_statements, start=1):
                row = status_by_table[(statement.source_db, statement.source_table)]
                table_key = (statement.source_db, statement.source_table)
                if table_key in failed_tables:
                    progress.progress(number / max(1, len(selected_statements)))
                    continue
                try:
                    cursor.execute(statement.sql)
                except Exception as exc:
                    row["결과"] = "실패"
                    row["메시지"] = str(exc)
                    failed_tables.add(table_key)
                row["실행 구문 수"] = int(row["실행 구문 수"]) + 1
                progress.progress(number / max(1, len(selected_statements)))
        conn.commit()
    progress.empty()
    return pd.DataFrame(status_by_table.values())


def render_header() -> None:
    st.markdown(
        """
        <style>
        .block-container { max-width: 1640px; padding-top: 1.4rem; padding-bottom: 2.5rem; }
        .tlp-hero { background: linear-gradient(115deg, #102A43 0%, #005EA8 48%, #00A6A6 100%); border-radius: 16px; color: white; padding: 28px 32px; text-align: center; box-shadow: 0 12px 28px rgba(0, 94, 168, 0.20); margin-bottom: 1.4rem; }
        .tlp-hero-title { font-size: 30px; font-weight: 750; letter-spacing: -0.6px; }
        .tlp-hero-subtitle { margin-top: 8px; font-size: 14px; color: #E6F6FF; }
        .tlp-hero-credit { margin-top: 12px; font-size: 12px; color: #C6F6F6; }
        </style>
        <div class="tlp-hero">
            <div class="tlp-hero-title">TLP DB 변경 내역</div>
            <div class="tlp-hero-subtitle">MSSQL 레이아웃 비교 · Oracle 스테이징 DDL 생성</div>
            <div class="tlp-hero-credit">⚙️ Created by ♡홍율파파♡</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_comparison() -> tuple[list[str], pd.DataFrame]:
    comparison = st.session_state.comparison
    if comparison is None:
        st.info("비교 기준일과 대상 기준일을 선택한 뒤 변경 내역 조회를 실행해 주십시오.", icon=":material/info:")
        return [], pd.DataFrame()
    table_changes: pd.DataFrame = comparison["table_changes"]
    column_changes: pd.DataFrame = comparison["column_changes"]
    timing = comparison.get("timing")
    if timing:
        st.caption(
            f"레이아웃 조회 {timing['load_seconds']:.1f}초 · 변경 분석 {timing['compare_seconds']:.1f}초 · 대상 행 {timing['row_count']:,}건"
        )
    if table_changes.empty:
        st.success("두 기준일 사이에 테이블 및 컬럼 변경이 없습니다.", icon=":material/check_circle:")
        return [], pd.DataFrame()
    summary = (
        table_changes.pivot_table(index="DB", columns="구분", values="테이블", aggfunc="count", fill_value=0)
        .reindex(columns=["신규", "삭제", "변경"], fill_value=0)
        .reset_index()
    )
    summary["합계"] = summary[["신규", "삭제", "변경"]].sum(axis=1)
    summary = pd.concat(
        [summary, pd.DataFrame([{"DB": "합계", "신규": summary["신규"].sum(), "삭제": summary["삭제"].sum(), "변경": summary["변경"].sum(), "합계": summary["합계"].sum()}])]
    )
    owners = sorted(table_changes["DB"].unique().tolist())
    summary_col, table_col = st.columns((0.85, 1.55))
    with summary_col:
        with st.container(border=True):
            st.markdown("#### :material/analytics: DB별 요약")
            st.caption("복수 선택 가능 · 미선택 또는 합계 선택 시 전체 DB")
            selection = st.dataframe(
                summary,
                hide_index=True,
                on_select="rerun",
                selection_mode="multi-row",
                key="db_summary",
                column_config=thousand_number_columns(summary),
                height=430,
            )
    selected_rows = selection.selection.rows
    selected_owners = [clean_text(summary.iloc[row]["DB"]) for row in selected_rows]
    if not selected_owners or "합계" in selected_owners:
        selected_owners = owners
    selected_tables = table_changes[table_changes["DB"].isin(selected_owners)].copy()
    selected_columns = column_changes[column_changes["DB"].isin(selected_owners)].copy()
    with table_col:
        with st.container(border=True):
            st.markdown("#### :material/table_chart: 테이블 변경 내역")
            st.caption("DDL을 만들 테이블 행을 하나 이상 선택해 주십시오.")
            table_selection = st.dataframe(
                selected_tables,
                hide_index=True,
                height=430,
                on_select="rerun",
                selection_mode="multi-row",
                key="table_selection",
                column_config=thousand_number_columns(selected_tables),
            )
    selected_rows = table_selection.selection.rows
    ddl_tables = selected_tables.iloc[selected_rows].copy() if selected_rows else pd.DataFrame(columns=selected_tables.columns)
    selected_table_keys = {(normalized(row["DB"]), normalized(row["테이블"])) for _, row in ddl_tables.iterrows()}
    column_display = selected_columns[
        selected_columns.apply(lambda row: (normalized(row["DB"]), normalized(row["테이블"])) in selected_table_keys, axis=1)
    ].copy() if selected_table_keys else selected_columns
    with st.container(border=True):
        st.markdown("#### :material/view_column: 컬럼 변경 내역")
        st.caption("테이블을 선택하면 선택 테이블의 컬럼 변경 내역만 표시합니다.")
        st.dataframe(column_display, hide_index=True, height=430, column_config=thousand_number_columns(column_display))
    return selected_owners, ddl_tables


def render_ddl_controls(selected_owners: list[str], selected_tables: pd.DataFrame, character_multiplier: int) -> None:
    if not selected_owners:
        return
    comparison = st.session_state.comparison
    if comparison is None:
        return
    selected_keys = tuple(sorted((normalized(row["DB"]), normalized(row["테이블"])) for _, row in selected_tables.iterrows()))
    with st.container(border=True):
        st.subheader(":material/code: Oracle DDL 생성")
        st.caption(f"선택 테이블: {len(selected_keys)}개")
        multiplier = st.number_input("문자형 길이 배수", min_value=1.0, max_value=4.0, value=float(character_multiplier), step=0.5, key="character_multiplier")
        if st.button("선택 테이블 DDL 생성", type="primary", icon=":material/code:", disabled=not selected_keys):
            try:
                st.session_state.ddl_artifact = build_ddl_artifact(
                    selected_tables,
                    comparison["before_layout"],
                    comparison["after_layout"],
                    DEFAULT_TARGET_OWNER,
                    DEFAULT_TABLESPACE,
                    DEFAULT_INDEX_TABLESPACE,
                    int(multiplier * 10) / 10,
                )
                st.session_state.ddl_logs = None
            except Exception as exc:
                st.error(str(exc))
    artifact: GeneratedDdl | None = st.session_state.ddl_artifact
    if artifact is None or artifact.source_keys != selected_keys:
        return
    st.success(f"DDL 생성 테이블: {artifact.table_count}개 · 생성 시각: {artifact.generated_at}", icon=":material/check_circle:")
    column_mapping = generated_column_mapping_grid(
        selected_tables,
        comparison["after_layout"],
        int(multiplier * 10) / 10,
    )
    structure_tab, comment_tab, column_mapping_tab, migration_mapping_tab = st.tabs(
        [":material/account_tree: 구조 DDL", ":material/comment: 코멘트 DDL", ":material/table_chart: 컬럼 매핑", ":material/sync_alt: 이관 매핑"]
    )
    with structure_tab:
        st.caption("구조 DDL에는 테이블·컬럼 코멘트가 함께 포함됩니다.")
        st.download_button(
            "구조 DDL 다운로드",
            data=artifact.text,
            file_name=f"MssqlOracleLayout_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql",
            mime="text/sql",
            icon=":material/download:",
        )
        if st.toggle("구조 DDL 미리보기", value=False, key="ddl_structure_preview"):
            render_ddl_preview(artifact.text)
    with comment_tab:
        st.caption(f"코멘트만 별도로 적용할 때 사용하는 DDL입니다. 테이블·컬럼 코멘트 {artifact.comment_count}건 · 컬럼 코멘트는 레이아웃의 ATTR 값을 사용합니다.")
        if artifact.comment_text:
            st.download_button(
                "코멘트 DDL 다운로드",
                data=artifact.comment_text,
                file_name=f"MssqlOracleLayout_Comments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql",
                mime="text/sql",
                icon=":material/download:",
            )
            if st.toggle("코멘트 DDL 미리보기", value=False, key="ddl_comment_preview"):
                render_ddl_preview(artifact.comment_text)
        else:
            st.info("생성할 코멘트가 없습니다. ENTITY 또는 ATTR 값이 있는지 확인해 주세요.")
    with column_mapping_tab:
        st.caption(f"DDL 생성 규칙으로 변환한 원천·타깃 컬럼 매핑 {len(column_mapping):,}건입니다.")
        st.download_button(
            "컬럼 매핑 엑셀 다운로드",
            data=column_mapping_excel_bytes(column_mapping),
            file_name=f"MssqlOracleLayout_ColumnMapping_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            icon=":material/download:",
        )
        st.dataframe(
            column_mapping,
            hide_index=True,
            height=460,
            column_config=thousand_number_columns(column_mapping),
        )
    with migration_mapping_tab:
        render_migration_mapping_sync(selected_tables, column_mapping)
    with st.expander(":material/play_arrow: Oracle DDL 실행", expanded=False):
        execution_mode = st.segmented_control(
            "실행 범위",
            ["전체 DDL", "코멘트 DDL만"],
            default="전체 DDL",
            key="execution_mode",
        )
        comment_only = execution_mode == "코멘트 DDL만"
        if comment_only:
            st.info("COMMENT ON TABLE 및 COMMENT ON COLUMN 구문만 실행합니다. DROP, CREATE, 인덱스, 제약은 실행하지 않습니다.")
        else:
            st.warning("변경·삭제 대상에는 DROP TABLE이 실행됩니다. 실행 실패 SQL은 계속 진행한 뒤 결과를 남깁니다.")
        execution_targets = generated_table_grid(artifact, comment_only=comment_only)
        if execution_targets.empty:
            st.info("선택한 실행 범위에 해당하는 DDL이 없습니다.")
            return
        st.caption("실행할 테이블 행을 선택해 주십시오.")
        execution_selection = st.dataframe(
            execution_targets,
            hide_index=True,
            on_select="rerun",
            selection_mode="multi-row",
            key="execution_table_selection",
            height=260,
            column_config=thousand_number_columns(execution_targets),
        )
        execution_rows = execution_selection.selection.rows
        execution_keys = {
            (normalized(execution_targets.iloc[row]["DB"]), normalized(execution_targets.iloc[row]["테이블"]))
            for row in execution_rows
        }
        confirmation_label = "생성된 COMMENT DDL 실행을 확인했습니다." if comment_only else "생성된 DDL의 DROP 및 CREATE 실행을 확인했습니다."
        confirmed = st.checkbox(confirmation_label, key="execute_confirm")
        button_label = "선택 테이블 COMMENT DDL 실행" if comment_only else "선택 테이블 Oracle DDL 실행"
        if st.button(button_label, type="primary", icon=":material/play_arrow:", disabled=not confirmed or not execution_keys):
            st.session_state.ddl_logs = execute_ddl(artifact, execution_keys, comment_only=comment_only)
    if st.session_state.ddl_logs is not None:
        st.subheader(":material/fact_check: DDL 실행 결과")
        st.dataframe(
            st.session_state.ddl_logs,
            hide_index=True,
            height=320,
            column_config=thousand_number_columns(st.session_state.ddl_logs),
        )


def main() -> None:
    st.set_page_config(page_title="TLP DB 변경 내역", page_icon=":material/difference:", layout="wide")
    init_state()
    render_header()
    layout_owner = sql_name(str(st.secrets["oracle"]["user"]))
    layout_table = DEFAULT_LAYOUT_TABLE
    try:
        snapshot_dates = fetch_snapshot_dates(layout_owner, layout_table)
    except Exception as exc:
        st.error(f"레이아웃 기준일을 조회할 수 없습니다: {exc}")
        st.stop()
    if len(snapshot_dates) < 2:
        st.warning("비교할 기준일이 두 개 이상 필요합니다.")
        st.stop()
    labels = [snapshot_label(value) for value in snapshot_dates]
    default_before_index = labels.index("2026-07-03") if "2026-07-03" in labels else max(0, len(labels) - 2)
    st.subheader(":material/date_range: 변경 내역 조회")
    with st.form("comparison_form", border=True):
        before_col, after_col, action_col = st.columns((1, 1, 0.7), vertical_alignment="bottom")
        with before_col:
            before_label = st.selectbox("비교 기준일", labels, index=default_before_index)
        with after_col:
            after_label = st.selectbox("대상 기준일", labels, index=len(labels) - 1)
        with action_col:
            submitted = st.form_submit_button("변경 내역 조회", type="primary", icon=":material/search:")
    if submitted:
        if before_label == after_label:
            st.error("서로 다른 두 기준일을 선택해 주십시오.")
        else:
            try:
                before_token = snapshot_dates[labels.index(before_label)]
                after_token = snapshot_dates[labels.index(after_label)]
                with st.status("변경 내역 조회를 시작합니다.", expanded=True) as status:
                    status.write("두 기준일의 레이아웃 데이터를 읽고 있습니다.")
                    load_started = perf_counter()
                    before_layout, after_layout = fetch_layout_pair(layout_owner, layout_table, before_token, after_token)
                    load_seconds = perf_counter() - load_started
                    status.write(f"레이아웃 조회 완료: {len(before_layout) + len(after_layout):,}건. 테이블·컬럼 변경을 분석하고 있습니다.")
                    compare_started = perf_counter()
                    table_changes, column_changes = compare_layouts(before_layout, after_layout)
                    compare_seconds = perf_counter() - compare_started
                    table_changes = add_table_catalog_metrics(table_changes)
                    status.update(
                        label=f"변경 내역 분석 완료: 테이블 {len(table_changes):,}건 · 컬럼 {len(column_changes):,}건",
                        state="complete",
                        expanded=False,
                    )
                st.session_state.comparison = {
                    "before_layout": before_layout,
                    "after_layout": after_layout,
                    "table_changes": table_changes,
                    "column_changes": column_changes,
                    "timing": {
                        "load_seconds": load_seconds,
                        "compare_seconds": compare_seconds,
                        "row_count": len(before_layout) + len(after_layout),
                    },
                }
                st.session_state.ddl_artifact = None
                st.session_state.ddl_logs = None
                st.session_state.selected_owner = None
            except Exception as exc:
                st.error(f"변경 내역 조회에 실패했습니다: {exc}")
    selected_owner, selected_tables = render_comparison()
    render_ddl_controls(selected_owner, selected_tables, 3)


if __name__ == "__main__":
    main()
